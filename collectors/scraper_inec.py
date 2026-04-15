"""
scraper_inec.py — Scraper del INEC (Instituto Nacional de Estadística y Censos)
=================================================================================
Portal:  https://www.ecuadorencifras.gob.ec
Datos:   Banco de Datos Abiertos INEC (ANDA), descarga directa de Excel/CSV
          - ENEMDU (Encuesta Nacional de Empleo, Desempleo y Subempleo)
          - IPC (Índice de Precios al Consumidor — Inflación)
          - Pobreza y desigualdad (GINI, FGT)
          - Censo 2022 de Población y Vivienda
          - Estadísticas vitales (nacimientos, defunciones)

Colección MongoDB: ecuador_intel.demografico.empleo
                   ecuador_intel.demografico.inflacion
                   ecuador_intel.demografico.pobreza
                   ecuador_intel.demografico.censo

Uso:
    python scraper_inec.py                     # Todo
    python scraper_inec.py --indicador empleo  # Solo empleo
    python scraper_inec.py --test              # Última versión de cada serie
"""

import argparse
import io
import json
import logging
import os
import re
import sys
import time
from datetime import datetime, timezone
from typing import Optional

import requests
from bs4 import BeautifulSoup
from pymongo import MongoClient, UpdateOne
from pymongo.errors import BulkWriteError

# ---------------------------------------------------------------------------
# Configuración
# ---------------------------------------------------------------------------

MONGO_URI = os.environ.get(
    "MONGO_URI",
    "mongodb+srv://scraperbot:GJqMqljz4GYBT0PU@cluster0.nz7wcxv.mongodb.net/"
    "?retryWrites=true&w=majority&appName=Cluster0",
)
DB_NAME      = "ecuador_intel"
COL_SYNC_LOG = "_sync_log"

INEC_BASE    = "https://www.ecuadorencifras.gob.ec"

# URLs de descarga directa de series estadísticas del INEC
SERIES = {
    "empleo": {
        "col": "demografico.empleo",
        "descripcion": "ENEMDU — Empleo, desempleo, subempleo trimestral",
        "urls": [
            "https://www.ecuadorencifras.gob.ec/documentos/web-inec/EMPLEO/2024/Serie-trimestral/Bases/"
            "BBDD_ENEMDU_trimestral_dic_24.csv",
        ],
        "pagina_busqueda": "https://www.ecuadorencifras.gob.ec/empleo-desempleo-y-subempleo/",
    },
    "inflacion": {
        "col": "demografico.inflacion",
        "descripcion": "IPC — Inflación mensual por división de gasto",
        "urls": [
            "https://www.ecuadorencifras.gob.ec/documentos/web-inec/Inflacion/2024/"
            "Diciembre-2024/1_ipc_serie_historica_2014_2024.xlsx",
        ],
        "pagina_busqueda": "https://www.ecuadorencifras.gob.ec/indice-de-precios-al-consumidor/",
    },
    "pobreza": {
        "col": "demografico.pobreza",
        "descripcion": "Pobreza y desigualdad — Incidencia, brechas, coeficiente GINI",
        "urls": [
            "https://www.ecuadorencifras.gob.ec/documentos/web-inec/PobrezayDesigualdad/"
            "Serie_historica_(2003-2023)/1_pobreza_desigualdad_serie_historica_2003_2023.xlsx",
        ],
        "pagina_busqueda": "https://www.ecuadorencifras.gob.ec/pobreza-y-desigualdad/",
    },
    "censo": {
        "col": "demografico.censo",
        "descripcion": "Censo 2022 — Resultados nacionales por provincia y cantón",
        "urls": [
            "https://www.ecuadorencifras.gob.ec/documentos/web-inec/Poblacion_y_Demografia/"
            "CPV_2022/Resultados_Nacionales/CPV_2022_Presentacion_Nacional.xlsx",
        ],
        "pagina_busqueda": "https://www.ecuadorencifras.gob.ec/censo-de-poblacion-y-vivienda/",
    },
}

HEADERS = {
    "User-Agent": "EcuaWatch-Bot/1.0 (INEC research; ecuawatch.org)"
}

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[logging.StreamHandler(sys.stdout)],
)
log = logging.getLogger("scraper_inec")

# ---------------------------------------------------------------------------
# Descarga y parseo
# ---------------------------------------------------------------------------

def descargar_archivo(url: str) -> Optional[bytes]:
    """Descarga binaria con reintentos."""
    for intento in range(3):
        try:
            r = requests.get(url, headers=HEADERS, timeout=60)
            r.raise_for_status()
            return r.content
        except requests.RequestException as e:
            log.warning("Intento %d fallido para %s: %s", intento + 1, url, e)
            time.sleep(3 ** intento)
    return None


def detectar_urls_reales(pagina_url: str) -> list[str]:
    """
    Parsea la página INEC buscando los enlaces de descarga más recientes.
    INEC usa WordPress; los archivos suelen estar en <a href="..."> dentro de
    divs con clase 'entry-content' o 'download-doc'.
    """
    urls = []
    try:
        r = requests.get(pagina_url, headers=HEADERS, timeout=30)
        soup = BeautifulSoup(r.text, "lxml")
        for a in soup.find_all("a", href=True):
            href = a["href"]
            # Buscar enlaces a archivos
            if any(ext in href.lower() for ext in [".csv", ".xlsx", ".xls", ".zip"]):
                if "ecuadorencifras" in href or href.startswith("/"):
                    if href.startswith("/"):
                        href = INEC_BASE + href
                    urls.append(href)
    except Exception as e:
        log.warning("No se pudo parsear %s: %s", pagina_url, e)
    return urls[:5]  # Máximo 5 archivos más recientes


def parsear_csv_inec(contenido: bytes, indicador: str) -> list[dict]:
    """
    Parsea CSV del INEC a lista de documentos normalizados.
    Detecta separador automáticamente.
    """
    import csv
    docs = []
    try:
        # Detectar encoding
        for enc in ("utf-8-sig", "latin-1", "iso-8859-1"):
            try:
                texto = contenido.decode(enc)
                break
            except UnicodeDecodeError:
                continue
        else:
            log.error("No se pudo decodificar el CSV de %s", indicador)
            return []

        # Detectar dialecto
        sample = texto[:2048]
        try:
            dialect = csv.Sniffer().sniff(sample)
        except csv.Error:
            dialect = csv.excel

        reader = csv.DictReader(io.StringIO(texto), dialect=dialect)
        for fila in reader:
            doc = {k.strip(): v.strip() for k, v in fila.items() if k}
            doc["_indicador"] = indicador
            doc["_fuente"]    = "INEC"
            doc["_ingestado"] = datetime.now(timezone.utc)
            docs.append(doc)

    except Exception as e:
        log.error("Error parseando CSV %s: %s", indicador, e)

    return docs


def parsear_excel_inec(contenido: bytes, indicador: str) -> list[dict]:
    """
    Parsea Excel del INEC. Intenta importar openpyxl; si no está disponible,
    guarda el binario como documento de referencia.
    """
    docs = []
    try:
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(contenido), read_only=True, data_only=True)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                continue
            cabeceras = [str(c).strip() if c else f"col_{i}" for i, c in enumerate(rows[0])]
            for fila in rows[1:]:
                if all(v is None for v in fila):
                    continue
                doc = dict(zip(cabeceras, fila))
                doc["_hoja"]      = sheet_name
                doc["_indicador"] = indicador
                doc["_fuente"]    = "INEC"
                doc["_ingestado"] = datetime.now(timezone.utc)
                docs.append(doc)

    except ImportError:
        log.warning("openpyxl no disponible. Guardando referencia binaria.")
        docs.append(
            {
                "_indicador":    indicador,
                "_fuente":       "INEC",
                "_tipo":         "excel_sin_parsear",
                "_tamaño_bytes": len(contenido),
                "_ingestado":    datetime.now(timezone.utc),
            }
        )
    except Exception as e:
        log.error("Error parseando Excel %s: %s", indicador, e)

    return docs


# ---------------------------------------------------------------------------
# MongoDB
# ---------------------------------------------------------------------------

def upsert_indicador(col, docs: list[dict]) -> int:
    """Inserta o actualiza documentos de series estadísticas."""
    if not docs:
        return 0
    # Para series sin ID natural, usamos insert en bulk
    try:
        col.insert_many(docs, ordered=False)
        return len(docs)
    except BulkWriteError as e:
        return e.details.get("nInserted", 0)


def registrar_sync_log(col_log, fuente: str, estado: str, detalle: dict):
    col_log.insert_one(
        {
            "fuente":    fuente,
            "estado":    estado,
            "detalle":   detalle,
            "timestamp": datetime.now(timezone.utc),
        }
    )


# ---------------------------------------------------------------------------
# Motor principal
# ---------------------------------------------------------------------------

def procesar_indicador(series_cfg: dict, nombre: str, db, modo_test: bool):
    """Descarga y guarda un indicador específico del INEC."""
    col     = db[series_cfg["col"]]
    urls    = series_cfg.get("urls", [])

    # Intentar detectar URLs más recientes desde la página oficial
    pagina = series_cfg.get("pagina_busqueda")
    if pagina and not modo_test:
        urls_detectadas = detectar_urls_reales(pagina)
        if urls_detectadas:
            log.info("URLs detectadas automáticamente para %s: %d", nombre, len(urls_detectadas))
            urls = urls_detectadas + urls  # Nuevas primero

    if not urls:
        log.warning("Sin URLs para %s", nombre)
        return 0

    total_insertados = 0
    for url in urls[:1 if modo_test else len(urls)]:
        log.info("  Descargando %s desde %s", nombre, url)
        contenido = descargar_archivo(url)
        if not contenido:
            log.error("  Fallo al descargar %s", url)
            continue

        # Parsear según extensión
        ext = url.split(".")[-1].lower()
        if ext == "csv":
            docs = parsear_csv_inec(contenido, nombre)
        elif ext in ("xlsx", "xls"):
            docs = parsear_excel_inec(contenido, nombre)
        else:
            log.warning("Formato no soportado: .%s para %s", ext, nombre)
            continue

        # Agregar URL de origen
        for d in docs:
            d["_url_origen"] = url

        if modo_test:
            docs = docs[:50]
            log.info("  [TEST] Limitado a 50 filas")

        n = upsert_indicador(col, docs)
        total_insertados += n
        log.info("  %s: %d documentos insertados desde %s", nombre, n, url)
        time.sleep(1)

    return total_insertados


def main(indicadores_filtro: list[str] = None, modo_test: bool = False):
    log.info("=== EcuaWatch · scraper_inec.py ===")

    client = MongoClient(MONGO_URI)
    db     = client[DB_NAME]
    col_log = db[COL_SYNC_LOG]

    series_a_procesar = {
        k: v for k, v in SERIES.items()
        if not indicadores_filtro or k in indicadores_filtro
    }

    resumen = {}
    for nombre, cfg in series_a_procesar.items():
        log.info("Procesando: %s — %s", nombre, cfg["descripcion"])
        n = procesar_indicador(cfg, nombre, db, modo_test)
        resumen[nombre] = n

    total = sum(resumen.values())
    log.info("Finalizado. Total insertados: %d | Desglose: %s", total, resumen)

    registrar_sync_log(
        col_log,
        fuente="inec",
        estado="completado",
        detalle={"resumen": resumen, "total": total},
    )

    client.close()


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="Scraper INEC — ecuadorencifras.gob.ec → MongoDB"
    )
    parser.add_argument(
        "--indicador",
        nargs="+",
        choices=list(SERIES.keys()),
        metavar="SERIE",
        help=f"Indicadores a procesar: {', '.join(SERIES.keys())}",
    )
    parser.add_argument(
        "--test",
        action="store_true",
        help="Modo prueba: solo último archivo, máximo 50 filas",
    )
    args = parser.parse_args()
    main(indicadores_filtro=args.indicador, modo_test=args.test)
