"""
scraper_contraloria.py — Scraper de la Contraloría General del Estado (CGE)
============================================================================
Portal:  https://www.contraloria.gob.ec
Datos:   Informes de auditoría, resoluciones, órdenes de trabajo,
         declaraciones patrimoniales y plan anual de control.

Colección MongoDB: ecuador_intel.fiscalizacion.informes_auditoria
                   ecuador_intel.fiscalizacion.resoluciones
                   ecuador_intel.fiscalizacion.ordenes_trabajo

Uso:
    python scraper_contraloria.py                      # Todo
    python scraper_contraloria.py --seccion informes   # Solo informes
    python scraper_contraloria.py --test               # Modo prueba
"""

import argparse
import csv
import io
import json
import logging
import os
import re
import sys
import time
from datetime import datetime, timezone
from typing import Optional
from urllib.parse import urljoin

import requests
from bs4 import BeautifulSoup
from pymongo import MongoClient, UpdateOne
from pymongo.errors import BulkWriteError

# ---------------------------------------------------------------------------
# Configuración
# ---------------------------------------------------------------------------

MONGO_URI = os.environ.get(
    "MONGO_URI",
    "mongodb+srv://scraperbot:GJqMqljz4GYBT0PU@cluster0.nz7wcxv.mongodb.net/"
    "?retryWrites=true&w=majority&appName=Cluster0",
)
DB_NAME      = "ecuador_intel"
COL_SYNC_LOG = "_sync_log"

CGE_BASE = "https://www.contraloria.gob.ec"

# Secciones de datos de la CGE
SECCIONES_CGE = {
    "informes": {
        "col":         "fiscalizacion.informes_auditoria",
        "descripcion": "Informes de auditoría aprobados por la CGE",
        "urls": [
            f"{CGE_BASE}/Informes/InformesAprobados",
            f"{CGE_BASE}/WFInformes.aspx",
        ],
        "tipo": "informes_auditoria",
    },
    "resoluciones": {
        "col":         "fiscalizacion.resoluciones",
        "descripcion": "Resoluciones de responsabilidades administrativas y civiles",
        "urls": [
            f"{CGE_BASE}/Consultas/Resoluciones",
            f"{CGE_BASE}/WFResoluciones.aspx",
        ],
        "tipo": "resoluciones",
    },
    "ordenes_trabajo": {
        "col":         "fiscalizacion.ordenes_trabajo",
        "descripcion": "Órdenes de trabajo de auditoría activas",
        "urls": [
            f"{CGE_BASE}/Consultas/OrdenesTrabajo",
            f"{CGE_BASE}/WFOrdenesTrabajo.aspx",
        ],
        "tipo": "ordenes_trabajo",
    },
    "datos_abiertos": {
        "col":         "fiscalizacion.datos_abiertos_cge",
        "descripcion": "Datasets publicados por la CGE en su portal de datos abiertos",
        "urls": [
            f"{CGE_BASE}/DatosAbiertos",
            f"{CGE_BASE}/Transparencia",
        ],
        "tipo": "datos_abiertos",
    },
}

HEADERS = {
    "User-Agent": "EcuaWatch-Bot/1.0 (CGE research; ecuawatch.org)",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    "Accept-Language": "es-EC,es;q=0.9,en;q=0.5",
}

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[logging.StreamHandler(sys.stdout)],
)
log = logging.getLogger("scraper_contraloria")

# ---------------------------------------------------------------------------
# HTTP helpers
# ---------------------------------------------------------------------------

def descargar(url: str, timeout: int = 60) -> Optional[requests.Response]:
    """Descarga con reintentos y backoff exponencial."""
    for intento in range(3):
        try:
            r = requests.get(url, headers=HEADERS, timeout=timeout, allow_redirects=True)
            r.raise_for_status()
            return r
        except requests.RequestException as e:
            log.warning("Intento %d fallido para %s: %s", intento + 1, url, e)
            time.sleep(3 ** intento)
    return None


def buscar_archivos_descargables(url: str) -> list[dict]:
    """
    Parsea una página de la CGE buscando enlaces a archivos descargables
    (CSV, Excel, PDF de datos, JSON).
    """
    archivos = []
    try:
        r = descargar(url)
        if not r:
            return archivos
        soup = BeautifulSoup(r.text, "lxml")
        extensiones = (".csv", ".xlsx", ".xls", ".json", ".ods", ".zip")

        for a in soup.find_all("a", href=True):
            href = a["href"]
            texto = a.get_text(strip=True)
            if any(ext in href.lower() for ext in extensiones):
                full_url = urljoin(url, href)
                archivos.append({
                    "url":    full_url,
                    "nombre": texto or href.split("/")[-1],
                    "ext":    href.split(".")[-1].lower().split("?")[0],
                })
    except Exception as e:
        log.warning("Error buscando archivos en %s: %s", url, e)
    return archivos


# ---------------------------------------------------------------------------
# Scraping de páginas HTML de la CGE
# ---------------------------------------------------------------------------

def scrape_tabla_html(url: str, tipo: str) -> list[dict]:
    """
    Muchas páginas de la CGE muestran datos en tablas HTML.
    Esta función extrae los datos de las tablas encontradas.
    """
    docs = []
    r = descargar(url)
    if not r:
        return docs

    soup = BeautifulSoup(r.text, "lxml")
    tablas = soup.find_all("table")

    for idx, tabla in enumerate(tablas):
        filas = tabla.find_all("tr")
        if len(filas) < 2:
            continue

        # Detectar cabeceras
        cabeceras = []
        primera_fila = filas[0]
        for th in primera_fila.find_all(["th", "td"]):
            texto = th.get_text(strip=True).replace("\n", " ")
            cabeceras.append(texto if texto else f"col_{len(cabeceras)}")

        if len(cabeceras) < 2:
            continue

        # Extraer datos
        for fila in filas[1:]:
            celdas = fila.find_all(["td", "th"])
            if not celdas:
                continue

            valores = [c.get_text(strip=True) for c in celdas]
            if all(not v for v in valores):
                continue

            doc = {}
            for j, val in enumerate(valores):
                if j < len(cabeceras):
                    doc[cabeceras[j]] = val
                else:
                    doc[f"col_{j}"] = val

            # Extraer enlaces dentro de las celdas (PDFs de informes)
            enlaces = []
            for c in celdas:
                for a in c.find_all("a", href=True):
                    href = a["href"]
                    if href and not href.startswith("#") and not href.startswith("javascript"):
                        enlaces.append(urljoin(url, href))
            if enlaces:
                doc["_enlaces_doc"] = enlaces

            doc["_tipo"]       = tipo
            doc["_fuente"]     = "CGE"
            doc["_tabla_idx"]  = idx
            doc["_url_origen"] = url
            doc["_ingestado"]  = datetime.now(timezone.utc)
            docs.append(doc)

    return docs


# ---------------------------------------------------------------------------
# Parseo Excel/CSV
# ---------------------------------------------------------------------------

def parsear_excel(contenido: bytes, nombre_seccion: str, url_origen: str) -> list[dict]:
    """Parsea archivo Excel de la CGE."""
    docs = []
    try:
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(contenido), read_only=True, data_only=True)

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            todas_filas = list(ws.iter_rows(values_only=True))
            if not todas_filas:
                continue

            cabeceras = None
            inicio_datos = 0
            for i, fila in enumerate(todas_filas):
                no_vacias = [c for c in fila if c is not None and str(c).strip()]
                if len(no_vacias) >= 3:
                    cabeceras = [
                        str(c).strip().replace("\n", " ") if c else f"col_{j}"
                        for j, c in enumerate(fila)
                    ]
                    inicio_datos = i + 1
                    break

            if not cabeceras:
                continue

            for fila in todas_filas[inicio_datos:]:
                if all(v is None for v in fila):
                    continue
                doc = dict(zip(cabeceras, fila))
                doc["_seccion"]    = nombre_seccion
                doc["_hoja"]       = sheet_name
                doc["_fuente"]     = "CGE"
                doc["_url_origen"] = url_origen
                doc["_ingestado"]  = datetime.now(timezone.utc)
                docs.append(doc)

    except ImportError:
        log.warning("openpyxl no disponible — guardando referencia")
        docs.append({
            "_seccion":      nombre_seccion,
            "_fuente":       "CGE",
            "_tipo":         "excel_sin_parsear",
            "_tamaño_bytes": len(contenido),
            "_url_origen":   url_origen,
            "_ingestado":    datetime.now(timezone.utc),
        })
    except Exception as e:
        log.error("Error parseando Excel CGE %s: %s", nombre_seccion, e)

    return docs


def parsear_csv(contenido: bytes, nombre_seccion: str, url_origen: str) -> list[dict]:
    """Parsea CSV con detección de encoding."""
    docs = []
    for enc in ("utf-8-sig", "latin-1", "cp1252"):
        try:
            texto = contenido.decode(enc)
            reader = csv.DictReader(io.StringIO(texto))
            for row in reader:
                row["_seccion"]    = nombre_seccion
                row["_fuente"]     = "CGE"
                row["_url_origen"] = url_origen
                row["_ingestado"]  = datetime.now(timezone.utc)
                docs.append(row)
            break
        except Exception:
            continue
    return docs


# ---------------------------------------------------------------------------
# CKAN complementario — datosabiertos.gob.ec
# ---------------------------------------------------------------------------

def buscar_ckan_contraloria(max_results: int = 5) -> list[dict]:
    """Busca datasets de la Contraloría en el portal CKAN."""
    recursos = []
    try:
        r = requests.get(
            "https://www.datosabiertos.gob.ec/api/3/action/package_search",
            params={"q": "contraloría auditoría", "rows": max_results},
            headers=HEADERS,
            timeout=30,
        )
        data = r.json()
        if data.get("success"):
            for pkg in data["result"]["results"]:
                for rec in pkg.get("resources", []):
                    fmt = (rec.get("format") or "").upper()
                    if fmt in ("CSV", "XLS", "XLSX", "ODS", "JSON"):
                        recursos.append({
                            "nombre":  rec.get("name") or pkg.get("title", "sin_nombre"),
                            "url":     rec.get("url"),
                            "formato": fmt,
                        })
    except Exception as e:
        log.warning("Error buscando en CKAN: %s", e)
    return recursos


# ---------------------------------------------------------------------------
# Driver por sección
# ---------------------------------------------------------------------------

def procesar_seccion(nombre: str, cfg: dict, db, modo_test: bool) -> int:
    col = db[cfg["col"]]
    col.create_index("_tipo")
    col.create_index("_ingestado")

    total = 0

    # 1. Scrape tablas HTML de las páginas de la CGE
    for url in cfg.get("urls", []):
        log.info("  Scrapeando tabla HTML: %s", url)
        docs = scrape_tabla_html(url, cfg["tipo"])
        if docs:
            if modo_test:
                docs = docs[:30]
            try:
                col.insert_many(docs, ordered=False)
                log.info("  Tabla HTML: %d documentos insertados desde %s", len(docs), url)
                total += len(docs)
            except BulkWriteError as e:
                n = e.details.get("nInserted", 0)
                log.warning("  BulkWriteError — %d insertados", n)
                total += n
        time.sleep(2)

    # 2. Buscar archivos descargables en las páginas
    for url in cfg.get("urls", []):
        archivos = buscar_archivos_descargables(url)
        for archivo in archivos[:3 if modo_test else 10]:
            log.info("  Descargando: %s (%s)", archivo["nombre"], archivo["ext"])
            r = descargar(archivo["url"])
            if not r:
                continue

            if archivo["ext"] in ("xls", "xlsx"):
                docs = parsear_excel(r.content, nombre, archivo["url"])
            elif archivo["ext"] == "csv":
                docs = parsear_csv(r.content, nombre, archivo["url"])
            elif archivo["ext"] == "json":
                try:
                    data = json.loads(r.content)
                    docs = data if isinstance(data, list) else [data]
                    for d in docs:
                        d["_seccion"]    = nombre
                        d["_fuente"]     = "CGE"
                        d["_url_origen"] = archivo["url"]
                        d["_ingestado"]  = datetime.now(timezone.utc)
                except Exception:
                    docs = []
            else:
                docs = [{
                    "_seccion":      nombre,
                    "_fuente":       "CGE",
                    "_tipo":         f"archivo_{archivo['ext']}",
                    "_tamaño_bytes": len(r.content),
                    "_url_origen":   archivo["url"],
                    "_ingestado":    datetime.now(timezone.utc),
                }]

            if modo_test:
                docs = docs[:20]

            if docs:
                try:
                    col.insert_many(docs, ordered=False)
                    log.info("  %d docs insertados de %s", len(docs), archivo["nombre"])
                    total += len(docs)
                except BulkWriteError as e:
                    n = e.details.get("nInserted", 0)
                    total += n

            time.sleep(1)

    # 3. CKAN complementario
    if nombre == "datos_abiertos":
        recursos_ckan = buscar_ckan_contraloria()
        for rec in recursos_ckan[:3]:
            log.info("  CKAN CGE: %s", rec["nombre"])
            r = descargar(rec["url"])
            if not r:
                continue
            if rec["formato"] in ("CSV",):
                docs = parsear_csv(r.content, nombre, rec["url"])
            elif rec["formato"] in ("XLS", "XLSX"):
                docs = parsear_excel(r.content, nombre, rec["url"])
            else:
                docs = []

            if docs:
                if modo_test:
                    docs = docs[:20]
                try:
                    col.insert_many(docs, ordered=False)
                    total += len(docs)
                except BulkWriteError as e:
                    total += e.details.get("nInserted", 0)

    return total


# ---------------------------------------------------------------------------
# Sync log
# ---------------------------------------------------------------------------

def registrar_sync_log(col_log, fuente: str, estado: str, detalle: dict):
    col_log.insert_one({
        "fuente":    fuente,
        "estado":    estado,
        "detalle":   detalle,
        "timestamp": datetime.now(timezone.utc),
    })


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main(secciones_filtro: list[str] = None, modo_test: bool = False):
    log.info("=== EcuaWatch · scraper_contraloria.py ===")

    client  = MongoClient(MONGO_URI)
    db      = client[DB_NAME]
    col_log = db[COL_SYNC_LOG]

    secciones = {
        k: v for k, v in SECCIONES_CGE.items()
        if not secciones_filtro or k in secciones_filtro
    }

    resumen = {}
    for nombre, cfg in secciones.items():
        log.info("Sección CGE: %s — %s", nombre, cfg["descripcion"])
        n = procesar_seccion(nombre, cfg, db, modo_test)
        resumen[nombre] = n

    total = sum(resumen.values())
    log.info("Finalizado CGE. Total: %d | %s", total, resumen)

    registrar_sync_log(
        col_log,
        fuente="contraloria",
        estado="completado",
        detalle={"resumen": resumen, "total": total},
    )
    client.close()


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="Scraper CGE — contraloria.gob.ec → MongoDB"
    )
    parser.add_argument(
        "--seccion",
        nargs="+",
        choices=list(SECCIONES_CGE.keys()),
        metavar="SECCION",
        help=f"Secciones a procesar: {', '.join(SECCIONES_CGE.keys())}",
    )
    parser.add_argument("--test", action="store_true", help="Modo prueba")
    args = parser.parse_args()
    main(secciones_filtro=args.seccion, modo_test=args.test)
