"""
scraper_minfin.py — Scraper del Ministerio de Economía y Finanzas (MEF)
========================================================================
Portal:  https://www.finanzas.gob.ec
Datos:   Ejecución del Presupuesto General del Estado (PGE)
           - Ingresos y gastos por sector
           - Deuda pública interna y externa
           - Transferencias a GADs
           - Proforma presupuestaria
         Portal complementario: datosabiertos.gob.ec (CKAN, filtro MEF)

Colección MongoDB: ecuador_intel.fiscal.presupuesto
                   ecuador_intel.fiscal.deuda_publica
                   ecuador_intel.fiscal.transferencias_gad

Uso:
    python scraper_minfin.py                  # Todo
    python scraper_minfin.py --seccion deuda  # Solo deuda pública
    python scraper_minfin.py --test           # Modo prueba
"""

import argparse
import csv
import io
import json
import logging
import os
import sys
import time
from datetime import datetime, timezone
from typing import Optional

import requests
from bs4 import BeautifulSoup
from pymongo import MongoClient, UpdateOne
from pymongo.errors import BulkWriteError

# ---------------------------------------------------------------------------
# Configuración
# ---------------------------------------------------------------------------

MONGO_URI = os.environ.get(
    "MONGO_URI",
    "mongodb+srv://scraperbot:GJqMqljz4GYBT0PU@cluster0.nz7wcxv.mongodb.net/"
    "?retryWrites=true&w=majority&appName=Cluster0",
)
DB_NAME      = "ecuador_intel"
COL_SYNC_LOG = "_sync_log"

# URLs del Ministerio de Finanzas
FINANZAS_BASE   = "https://www.finanzas.gob.ec"
CKAN_API        = "https://www.datosabiertos.gob.ec/api/3/action"

# Secciones de datos del MEF
SECCIONES_MEF = {
    "presupuesto": {
        "col":          "fiscal.presupuesto",
        "descripcion":  "Ejecución presupuestaria del PGE — ingresos y gastos",
        "ckan_query":   "ejecución presupuestaria finanzas",
        "urls_directas": [
            f"{FINANZAS_BASE}/wp-content/uploads/downloads/estadisticas/"
            "ejecucion_presupuestaria.xlsx",
        ],
        "paginas_busqueda": [
            f"{FINANZAS_BASE}/informacion-presupuestaria/",
        ],
    },
    "deuda": {
        "col":          "fiscal.deuda_publica",
        "descripcion":  "Deuda pública interna y externa del Ecuador",
        "ckan_query":   "deuda pública Ecuador finanzas",
        "urls_directas": [
            f"{FINANZAS_BASE}/wp-content/uploads/downloads/estadisticas/"
            "deuda_publica.xlsx",
        ],
        "paginas_busqueda": [
            f"{FINANZAS_BASE}/deuda-publica/",
        ],
    },
    "transferencias_gad": {
        "col":          "fiscal.transferencias_gad",
        "descripcion":  "Transferencias del gobierno central a GADs (municipios/prefecturas)",
        "ckan_query":   "transferencias GAD municipios finanzas",
        "urls_directas": [],
        "paginas_busqueda": [
            f"{FINANZAS_BASE}/transferencias-a-gad/",
        ],
    },
}

HEADERS = {
    "User-Agent": "EcuaWatch-Bot/1.0 (MEF research; ecuawatch.org)"
}

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[logging.StreamHandler(sys.stdout)],
)
log = logging.getLogger("scraper_minfin")

# ---------------------------------------------------------------------------
# HTTP helpers
# ---------------------------------------------------------------------------

def descargar(url: str, timeout: int = 60) -> Optional[bytes]:
    """Descarga con reintentos y backoff exponencial."""
    for intento in range(3):
        try:
            r = requests.get(url, headers=HEADERS, timeout=timeout, allow_redirects=True)
            r.raise_for_status()
            return r.content
        except requests.RequestException as e:
            log.warning("Intento %d fallido para %s: %s", intento + 1, url, e)
            time.sleep(3 ** intento)
    return None


def buscar_archivos_en_pagina(url_base: str, extensiones=(".xlsx", ".xls", ".csv", ".ods")) -> list[str]:
    """Parsea una página del MEF buscando enlaces directos a archivos de datos."""
    urls = []
    try:
        r = requests.get(url_base, headers=HEADERS, timeout=30)
        soup = BeautifulSoup(r.text, "lxml")
        for a in soup.find_all("a", href=True):
            href = a["href"]
            if any(ext in href.lower() for ext in extensiones):
                if href.startswith("http"):
                    urls.append(href)
                elif href.startswith("/"):
                    urls.append(f"{FINANZAS_BASE}{href}")
                else:
                    base = "/".join(url_base.split("/")[:-1])
                    urls.append(f"{base}/{href}")
    except Exception as e:
        log.warning("Error parseando página MEF %s: %s", url_base, e)
    urls.sort(reverse=True)
    return urls[:5]


# ---------------------------------------------------------------------------
# CKAN — datosabiertos.gob.ec
# ---------------------------------------------------------------------------

def buscar_datasets_ckan(query: str, max_results: int = 5) -> list[dict]:
    """
    Busca datasets en el portal CKAN de Datos Abiertos de Ecuador.
    Retorna lista de URLs de recursos descargables.
    """
    recursos = []
    try:
        r = requests.get(
            f"{CKAN_API}/package_search",
            params={"q": query, "rows": max_results},
            headers=HEADERS,
            timeout=30,
        )
        data = r.json()
        if data.get("success"):
            for pkg in data["result"]["results"]:
                for rec in pkg.get("resources", []):
                    fmt = (rec.get("format") or "").upper()
                    if fmt in ("CSV", "XLS", "XLSX", "ODS", "JSON"):
                        recursos.append({
                            "nombre":  rec.get("name") or pkg.get("title", "sin_nombre"),
                            "url":     rec.get("url"),
                            "formato": fmt,
                            "paquete": pkg.get("name"),
                            "org":     pkg.get("organization", {}).get("title", ""),
                        })
    except Exception as e:
        log.warning("Error buscando en CKAN: %s", e)
    return recursos


# ---------------------------------------------------------------------------
# Parseo Excel/CSV
# ---------------------------------------------------------------------------

def parsear_excel(contenido: bytes, nombre_serie: str, url_origen: str) -> list[dict]:
    """Parsea archivo Excel del MEF con detección automática de cabeceras."""
    docs = []
    try:
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(contenido), read_only=True, data_only=True)

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            todas_filas = list(ws.iter_rows(values_only=True))
            if not todas_filas:
                continue

            cabeceras = None
            inicio_datos = 0
            for i, fila in enumerate(todas_filas):
                no_vacias = [c for c in fila if c is not None and str(c).strip()]
                if len(no_vacias) >= 3:
                    cabeceras = [
                        str(c).strip().replace("\n", " ") if c else f"col_{j}"
                        for j, c in enumerate(fila)
                    ]
                    inicio_datos = i + 1
                    break

            if not cabeceras:
                continue

            for fila in todas_filas[inicio_datos:]:
                if all(v is None for v in fila):
                    continue
                doc = dict(zip(cabeceras, fila))
                doc["_seccion"]    = nombre_serie
                doc["_hoja"]       = sheet_name
                doc["_fuente"]     = "MEF"
                doc["_url_origen"] = url_origen
                doc["_ingestado"]  = datetime.now(timezone.utc)
                docs.append(doc)

    except ImportError:
        log.warning("openpyxl no disponible — guardando referencia de %s", nombre_serie)
        docs.append({
            "_seccion":        nombre_serie,
            "_fuente":         "MEF",
            "_tipo":           "excel_sin_parsear",
            "_tamaño_bytes":   len(contenido),
            "_url_origen":     url_origen,
            "_ingestado":      datetime.now(timezone.utc),
        })
    except Exception as e:
        log.error("Error parseando Excel MEF %s: %s", nombre_serie, e)

    return docs


def parsear_csv(contenido: bytes, nombre_serie: str, url_origen: str) -> list[dict]:
    """Parsea CSV con detección de encoding."""
    docs = []
    for enc in ("utf-8-sig", "latin-1", "cp1252"):
        try:
            texto = contenido.decode(enc)
            reader = csv.DictReader(io.StringIO(texto))
            for row in reader:
                row["_seccion"]    = nombre_serie
                row["_fuente"]     = "MEF"
                row["_url_origen"] = url_origen
                row["_ingestado"]  = datetime.now(timezone.utc)
                docs.append(row)
            break
        except Exception:
            continue
    return docs


# ---------------------------------------------------------------------------
# Driver por sección
# ---------------------------------------------------------------------------

def procesar_seccion(nombre: str, cfg: dict, db, modo_test: bool) -> int:
    col = db[cfg["col"]]
    col.create_index("_seccion")
    col.create_index("_ingestado")

    urls = list(cfg.get("urls_directas", []))

    # Buscar archivos adicionales en las páginas del MEF
    for pagina in cfg.get("paginas_busqueda", []):
        extras = buscar_archivos_en_pagina(pagina)
        if extras:
            urls = extras + urls
            log.info("  Encontrados %d archivos en %s", len(extras), pagina)

    # Buscar en CKAN
    ckan_query = cfg.get("ckan_query", "")
    if ckan_query:
        recursos_ckan = buscar_datasets_ckan(ckan_query, max_results=3)
        for rec in recursos_ckan:
            if rec["url"] and rec["url"] not in urls:
                urls.append(rec["url"])
                log.info("  CKAN recurso: %s (%s)", rec["nombre"], rec["formato"])

    # Deduplicar
    urls = list(dict.fromkeys(urls))

    if not urls:
        log.warning("Sin URLs para la sección %s", nombre)
        return 0

    total = 0
    for url in urls[:2 if modo_test else 5]:
        log.info("  Descargando %s → %s", nombre, url)
        contenido = descargar(url)
        if not contenido:
            log.error("  Falló descarga: %s", url)
            continue

        ext = url.split("?")[0].split(".")[-1].lower()
        if ext in ("xls", "xlsx"):
            docs = parsear_excel(contenido, nombre, url)
        elif ext in ("csv", "ods"):
            docs = parsear_csv(contenido, nombre, url)
        elif ext == "json":
            try:
                data = json.loads(contenido)
                if isinstance(data, list):
                    docs = data
                elif isinstance(data, dict) and "data" in data:
                    docs = data["data"] if isinstance(data["data"], list) else [data]
                else:
                    docs = [data]
                for d in docs:
                    d["_seccion"]    = nombre
                    d["_fuente"]     = "MEF"
                    d["_url_origen"] = url
                    d["_ingestado"]  = datetime.now(timezone.utc)
            except Exception as e:
                log.error("Error parseando JSON: %s", e)
                docs = []
        else:
            # Guardar como referencia
            docs = [{
                "_seccion":      nombre,
                "_fuente":       "MEF",
                "_tipo":         f"archivo_{ext}",
                "_tamaño_bytes": len(contenido),
                "_url_origen":   url,
                "_ingestado":    datetime.now(timezone.utc),
            }]

        if modo_test:
            docs = docs[:30]

        if docs:
            try:
                col.insert_many(docs, ordered=False)
                log.info("  %s: %d documentos insertados", nombre, len(docs))
                total += len(docs)
            except BulkWriteError as e:
                n = e.details.get("nInserted", 0)
                log.warning("  BulkWriteError — %d insertados", n)
                total += n

        time.sleep(1)

    return total


# ---------------------------------------------------------------------------
# Sync log
# ---------------------------------------------------------------------------

def registrar_sync_log(col_log, fuente: str, estado: str, detalle: dict):
    col_log.insert_one({
        "fuente":    fuente,
        "estado":    estado,
        "detalle":   detalle,
        "timestamp": datetime.now(timezone.utc),
    })


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main(secciones_filtro: list[str] = None, modo_test: bool = False):
    log.info("=== EcuaWatch · scraper_minfin.py ===")

    client  = MongoClient(MONGO_URI)
    db      = client[DB_NAME]
    col_log = db[COL_SYNC_LOG]

    secciones = {
        k: v for k, v in SECCIONES_MEF.items()
        if not secciones_filtro or k in secciones_filtro
    }

    resumen = {}
    for nombre, cfg in secciones.items():
        log.info("Sección MEF: %s — %s", nombre, cfg["descripcion"])
        n = procesar_seccion(nombre, cfg, db, modo_test)
        resumen[nombre] = n

    total = sum(resumen.values())
    log.info("Finalizado MEF. Total: %d | %s", total, resumen)

    registrar_sync_log(
        col_log,
        fuente="minfin",
        estado="completado",
        detalle={"resumen": resumen, "total": total},
    )
    client.close()


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="Scraper MEF — finanzas.gob.ec → MongoDB"
    )
    parser.add_argument(
        "--seccion",
        nargs="+",
        choices=list(SECCIONES_MEF.keys()),
        metavar="SECCION",
        help=f"Secciones a procesar: {', '.join(SECCIONES_MEF.keys())}",
    )
    parser.add_argument("--test", action="store_true", help="Modo prueba")
    args = parser.parse_args()
    main(secciones_filtro=args.seccion, modo_test=args.test)
