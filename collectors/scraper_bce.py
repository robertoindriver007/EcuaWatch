"""
scraper_bce.py — Scraper del Banco Central del Ecuador (BCE)
=============================================================
Portal:  https://www.bce.fin.ec
Datos:   Series macroeconómicas históricas
          - PIB trimestral (nominal y real)
          - Inflación (IPC, IPP, DAP)
          - Balanza de pagos y comercio exterior
          - Liquidez total y reservas internacionales
          - Tasas de interés activas y pasivas
          - Remesas del exterior

Colección MongoDB: ecuador_intel.economico.indicadores_bce
                   ecuador_intel.economico.comercio_exterior
                   ecuador_intel.economico.tasas_interes

Uso:
    python scraper_bce.py              # Todos los indicadores
    python scraper_bce.py --serie pib  # Solo PIB
    python scraper_bce.py --test       # Modo prueba
"""

import argparse
import io
import logging
import os
import sys
import time
from datetime import datetime, timezone
from typing import Optional

import requests
from bs4 import BeautifulSoup
from pymongo import MongoClient, UpdateOne
from pymongo.errors import BulkWriteError

# ---------------------------------------------------------------------------
# Configuración
# ---------------------------------------------------------------------------

MONGO_URI = os.environ.get(
    "MONGO_URI",
    "mongodb+srv://scraperbot:GJqMqljz4GYBT0PU@cluster0.nz7wcxv.mongodb.net/"
    "?retryWrites=true&w=majority&appName=Cluster0",
)
DB_NAME      = "ecuador_intel"
COL_SYNC_LOG = "_sync_log"

BCE_BASE = "https://contenido.bce.fin.ec"

# Series del BCE con URLs de descarga directa (Excel/CSV de información estadística)
SERIES_BCE = {
    "pib": {
        "col":          "economico.pib",
        "descripcion":  "PIB trimestral — Precios corrientes y constantes, por industria",
        "urls": [
            f"{BCE_BASE}/documentos/Estadisticas/SectorReal/CuentasNacionales/"
            "Trimestrales/CartasTrimestralMacro/indice.htm",
            "https://www.bce.fin.ec/index.php/component/k2/item/download/203_"
            "be48e3024b29e9d9a3cc6e07f4d0ccc2",
        ],
        "tipo_busqueda": "cuentas_nacionales",
    },
    "inflacion": {
        "col":          "economico.inflacion_bce",
        "descripcion":  "Evolución mensual de inflación (IPC, IPP, deflactores)",
        "urls": [
            f"{BCE_BASE}/documentos/Estadisticas/SectorReal/Precios/IndicePrecios/"
            "Evolucion_IPC.xls",
        ],
        "tipo_busqueda": "precios",
    },
    "comercio_exterior": {
        "col":          "economico.comercio_exterior",
        "descripcion":  "Exportaciones e importaciones por grupos de productos y países",
        "urls": [
            f"{BCE_BASE}/documentos/Estadisticas/SectorExterno/BalanzaPagos/"
            "exportaciones_2024.xlsx",
            f"{BCE_BASE}/documentos/Estadisticas/SectorExterno/BalanzaPagos/"
            "importaciones_2024.xlsx",
        ],
        "tipo_busqueda": "sector_externo",
    },
    "tasas_interes": {
        "col":          "economico.tasas_interes",
        "descripcion":  "Tasas de interés activas y pasivas del sistema financiero",
        "urls": [
            f"{BCE_BASE}/documentos/Estadisticas/SectorMonFin/TasasInteres/"
            "Indice.htm",
        ],
        "tipo_busqueda": "tasas",
    },
    "reservas": {
        "col":          "economico.reservas_internacionales",
        "descripcion":  "Reservas Internacionales de Libre Disponibilidad (RILD)",
        "urls": [
            f"{BCE_BASE}/documentos/Estadisticas/SectorMonFin/ReservasInternacionales/"
            "Indice.htm",
        ],
        "tipo_busqueda": "reservas",
    },
    "remesas": {
        "col":          "economico.remesas",
        "descripcion":  "Remesas recibidas del exterior por trimestre y país de origen",
        "urls": [
            f"{BCE_BASE}/documentos/Estadisticas/SectorExterno/BalanzaPagos/"
            "Remesas/remesas_2024.xlsx",
        ],
        "tipo_busqueda": "remesas",
    },
}

HEADERS = {
    "User-Agent": "EcuaWatch-Bot/1.0 (BCE research; ecuawatch.org)"
}

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[logging.StreamHandler(sys.stdout)],
)
log = logging.getLogger("scraper_bce")

# ---------------------------------------------------------------------------
# HTTP helpers
# ---------------------------------------------------------------------------

def descargar(url: str, timeout: int = 60) -> Optional[bytes]:
    for intento in range(3):
        try:
            r = requests.get(url, headers=HEADERS, timeout=timeout)
            r.raise_for_status()
            return r.content
        except requests.RequestException as e:
            log.warning("Intento %d fallido para %s: %s", intento + 1, url, e)
            time.sleep(3 ** intento)
    return None


def buscar_excel_en_pagina(url_base: str) -> list[str]:
    """
    Parsea una página del BCE buscando enlaces directos a Excel/CSV.
    El BCE usa tablas de índice con años.
    """
    urls = []
    try:
        r = requests.get(url_base, headers=HEADERS, timeout=30)
        soup = BeautifulSoup(r.text, "lxml")
        base = "/".join(url_base.split("/")[:-1])
        for a in soup.find_all("a", href=True):
            href = a["href"]
            if any(ext in href.lower() for ext in [".xls", ".xlsx", ".csv"]):
                if href.startswith("http"):
                    urls.append(href)
                elif href.startswith("/"):
                    urls.append(f"https://contenido.bce.fin.ec{href}")
                else:
                    urls.append(f"{base}/{href}")
    except Exception as e:
        log.warning("Error parseando página BCE %s: %s", url_base, e)
    # Priorizar archivos más recientes (mayor año en URL)
    urls.sort(reverse=True)
    return urls[:3]


# ---------------------------------------------------------------------------
# Parseo Excel — BCE
# ---------------------------------------------------------------------------

def parsear_excel_bce(contenido: bytes, serie: str, url_origen: str) -> list[dict]:
    """
    Parsea el Excel del BCE. Los archivos del BCE tienen formatos variados:
    algunos son tablas simples, otros tienen metadatos en las primeras filas.
    Intentamos detectar automáticamente dónde empiezan los datos.
    """
    docs = []
    try:
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(contenido), read_only=True, data_only=True)

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            todas_filas = list(ws.iter_rows(values_only=True))
            if not todas_filas:
                continue

            # Buscar la fila de cabecera (primera fila con >3 celdas no vacías)
            inicio_datos = 0
            cabeceras = None
            for i, fila in enumerate(todas_filas):
                no_vacias = [c for c in fila if c is not None and str(c).strip()]
                if len(no_vacias) >= 3:
                    cabeceras = [
                        str(c).strip().replace("\n", " ") if c else f"col_{j}"
                        for j, c in enumerate(fila)
                    ]
                    inicio_datos = i + 1
                    break

            if not cabeceras:
                continue

            for fila in todas_filas[inicio_datos:]:
                if all(v is None for v in fila):
                    continue
                doc = dict(zip(cabeceras, fila))
                doc["_serie"]      = serie
                doc["_hoja"]       = sheet_name
                doc["_fuente"]     = "BCE"
                doc["_url_origen"] = url_origen
                doc["_ingestado"]  = datetime.now(timezone.utc)
                docs.append(doc)

    except ImportError:
        log.warning("openpyxl no disponible — guardando referencia de %s", serie)
        docs.append(
            {
                "_serie":          serie,
                "_fuente":         "BCE",
                "_tipo":           "excel_sin_parsear",
                "_tamaño_bytes":   len(contenido),
                "_url_origen":     url_origen,
                "_ingestado":      datetime.now(timezone.utc),
            }
        )
    except Exception as e:
        log.error("Error parseando Excel BCE %s: %s", serie, e)

    return docs


# ---------------------------------------------------------------------------
# Driver indicador BCE
# ---------------------------------------------------------------------------

def procesar_serie(nombre: str, cfg: dict, db, modo_test: bool) -> int:
    col = db[cfg["col"]]
    # Índices
    col.create_index("_serie")
    col.create_index("_ingestado")

    urls = list(cfg.get("urls", []))

    # Complementar con búsqueda dinámica en páginas índice
    for url in urls[:]:
        if url.endswith(".htm") or url.endswith(".html"):
            extras = buscar_excel_en_pagina(url)
            urls = extras + [u for u in urls if not u.endswith((".htm", ".html"))]
            break

    if not urls:
        log.warning("Sin URLs para la serie %s", nombre)
        return 0

    total = 0
    for url in urls[:1 if modo_test else 3]:
        log.info("  Descargando %s → %s", nombre, url)
        contenido = descargar(url)
        if not contenido:
            log.error("  Falló descarga: %s", url)
            continue

        ext = url.split("?")[0].split(".")[-1].lower()
        if ext in ("xls", "xlsx"):
            docs = parsear_excel_bce(contenido, nombre, url)
        elif ext == "csv":
            import csv
            rows = []
            for enc in ("utf-8-sig", "latin-1"):
                try:
                    texto = contenido.decode(enc)
                    reader = csv.DictReader(io.StringIO(texto))
                    rows = list(reader)
                    break
                except Exception:
                    pass
            docs = []
            for row in rows:
                row["_serie"]      = nombre
                row["_fuente"]     = "BCE"
                row["_url_origen"] = url
                row["_ingestado"]  = datetime.now(timezone.utc)
                docs.append(row)
        else:
            log.warning("Formato no manejado: .%s — %s", ext, url)
            continue

        if modo_test:
            docs = docs[:30]

        if docs:
            try:
                col.insert_many(docs, ordered=False)
                log.info("  %s: %d documentos insertados", nombre, len(docs))
                total += len(docs)
            except BulkWriteError as e:
                n = e.details.get("nInserted", 0)
                log.warning("  BulkWriteError — %d insertados", n)
                total += n

        time.sleep(1)

    return total


# ---------------------------------------------------------------------------
# Sync log
# ---------------------------------------------------------------------------

def registrar_sync_log(col_log, fuente: str, estado: str, detalle: dict):
    col_log.insert_one(
        {
            "fuente":    fuente,
            "estado":    estado,
            "detalle":   detalle,
            "timestamp": datetime.now(timezone.utc),
        }
    )


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main(series_filtro: list[str] = None, modo_test: bool = False):
    log.info("=== EcuaWatch · scraper_bce.py ===")

    client  = MongoClient(MONGO_URI)
    db      = client[DB_NAME]
    col_log = db[COL_SYNC_LOG]

    series = {
        k: v for k, v in SERIES_BCE.items()
        if not series_filtro or k in series_filtro
    }

    resumen = {}
    for nombre, cfg in series.items():
        log.info("Serie BCE: %s — %s", nombre, cfg["descripcion"])
        n = procesar_serie(nombre, cfg, db, modo_test)
        resumen[nombre] = n

    total = sum(resumen.values())
    log.info("Finalizado BCE. Total: %d | %s", total, resumen)

    registrar_sync_log(
        col_log,
        fuente="bce",
        estado="completado",
        detalle={"resumen": resumen, "total": total},
    )
    client.close()


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="Scraper BCE — bce.fin.ec → Mongolia"
    )
    parser.add_argument(
        "--serie",
        nargs="+",
        choices=list(SERIES_BCE.keys()),
        metavar="SERIE",
        help=f"Series a procesar: {', '.join(SERIES_BCE.keys())}",
    )
    parser.add_argument("--test", action="store_true", help="Modo prueba")
    args = parser.parse_args()
    main(series_filtro=args.serie, modo_test=args.test)
