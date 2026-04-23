"""
scraper_sri.py — Scraper del Servicio de Rentas Internas (SRI)
===============================================================
Portal:  https://www.sri.gob.ec
Datos:   Estadísticas de recaudación tributaria
           - Recaudación mensual por tipo de impuesto
           - Recaudación por provincia
           - Catastro de contribuyentes (RUC)
           - Devoluciones de IVA
           - Estadísticas de comercio exterior (importaciones/exportaciones)
         Portal datos abiertos: sri.gob.ec/datos-abiertos

Colección MongoDB: ecuador_intel.tributario.recaudacion
                   ecuador_intel.tributario.recaudacion_provincial
                   ecuador_intel.tributario.catastro_ruc
                   ecuador_intel.tributario.devoluciones

Uso:
    python scraper_sri.py                          # Todo
    python scraper_sri.py --seccion recaudacion    # Solo recaudación
    python scraper_sri.py --test                   # Modo prueba
"""

import argparse
import csv
import io
import json
import logging
import os
import re
import sys
import time
from datetime import datetime, timezone
from typing import Optional
from urllib.parse import urljoin

import requests
from bs4 import BeautifulSoup
from pymongo import MongoClient, UpdateOne
from pymongo.errors import BulkWriteError

# ---------------------------------------------------------------------------
# Configuración
# ---------------------------------------------------------------------------

MONGO_URI = os.environ.get(
    "MONGO_URI",
    "mongodb+srv://scraperbot:GJqMqljz4GYBT0PU@cluster0.nz7wcxv.mongodb.net/"
    "?retryWrites=true&w=majority&appName=Cluster0",
)
DB_NAME      = "ecuador_intel"
COL_SYNC_LOG = "_sync_log"

SRI_BASE         = "https://www.sri.gob.ec"
SRI_DATOS        = f"{SRI_BASE}/datos-abiertos"
SRI_ESTADISTICAS = f"{SRI_BASE}/estadisticas-generales-de-recaudacion-702"
CKAN_API         = "https://www.datosabiertos.gob.ec/api/3/action"

# Secciones de datos del SRI
SECCIONES_SRI = {
    "recaudacion": {
        "col":         "tributario.recaudacion",
        "descripcion": "Recaudación tributaria mensual por tipo de impuesto (IVA, IR, ICE, etc.)",
        "urls": [
            SRI_ESTADISTICAS,
            f"{SRI_BASE}/estadisticas-generales-de-recaudacion",
            f"{SRI_BASE}/web/guest/estadisticas-generales-de-recaudacion",
        ],
        "ckan_query": "recaudación tributaria SRI impuestos",
        "tipo":       "recaudacion_nacional",
    },
    "recaudacion_provincial": {
        "col":         "tributario.recaudacion_provincial",
        "descripcion": "Recaudación tributaria desglosada por provincia y cantón",
        "urls": [
            f"{SRI_BASE}/estadisticas-por-provincia",
            f"{SRI_BASE}/web/guest/estadisticas-generales-de-recaudacion-702",
        ],
        "ckan_query": "recaudación provincia SRI",
        "tipo":       "recaudacion_provincial",
    },
    "catastro_ruc": {
        "col":         "tributario.catastro_ruc",
        "descripcion": "Estadísticas del catastro de contribuyentes (RUC activos/pasivos)",
        "urls": [
            f"{SRI_BASE}/catastros",
            f"{SRI_BASE}/web/guest/catastros",
        ],
        "ckan_query": "catastro RUC contribuyentes SRI",
        "tipo":       "catastro",
    },
    "devoluciones": {
        "col":         "tributario.devoluciones",
        "descripcion": "Devoluciones de IVA, Impuesto a la Renta, y otros",
        "urls": [
            f"{SRI_BASE}/devoluciones-de-impuestos",
        ],
        "ckan_query": "devoluciones IVA SRI Ecuador",
        "tipo":       "devoluciones",
    },
    "datos_abiertos_sri": {
        "col":         "tributario.datos_abiertos_sri",
        "descripcion": "Todos los datasets publicados por el SRI en su portal de datos abiertos",
        "urls": [
            SRI_DATOS,
            f"{SRI_BASE}/datos-abiertos-702",
        ],
        "ckan_query": "SRI servicio rentas internas",
        "tipo":       "datos_abiertos",
    },
}

HEADERS = {
    "User-Agent": "EcuaWatch-Bot/1.0 (SRI research; ecuawatch.org)",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    "Accept-Language": "es-EC,es;q=0.9,en;q=0.5",
}

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[logging.StreamHandler(sys.stdout)],
)
log = logging.getLogger("scraper_sri")

# ---------------------------------------------------------------------------
# HTTP helpers
# ---------------------------------------------------------------------------

def descargar(url: str, timeout: int = 60) -> Optional[requests.Response]:
    """Descarga con reintentos y backoff exponencial."""
    for intento in range(3):
        try:
            r = requests.get(url, headers=HEADERS, timeout=timeout, allow_redirects=True)
            r.raise_for_status()
            return r
        except requests.RequestException as e:
            log.warning("Intento %d fallido para %s: %s", intento + 1, url, e)
            time.sleep(3 ** intento)
    return None


def buscar_archivos_en_pagina(url: str) -> list[dict]:
    """
    Parsea una página del SRI buscando enlaces a archivos descargables.
    El SRI suele publicar sus estadísticas en Excel.
    """
    archivos = []
    try:
        r = descargar(url)
        if not r:
            return archivos

        soup = BeautifulSoup(r.text, "lxml")
        extensiones = (".csv", ".xlsx", ".xls", ".json", ".ods", ".zip", ".pdf")

        for a in soup.find_all("a", href=True):
            href = a["href"]
            texto = a.get_text(strip=True)

            # Priorizar archivos de datos, no PDFs genéricos
            if any(ext in href.lower() for ext in extensiones):
                full_url = urljoin(url, href)
                ext = href.split("?")[0].split(".")[-1].lower()

                # Excluir PDFs que no parecen ser datos estadísticos
                if ext == "pdf":
                    # Solo incluir PDFs con palabras clave de datos
                    keywords = ["estadistic", "recaudac", "catastro", "devolucion", "dato"]
                    if not any(kw in texto.lower() or kw in href.lower() for kw in keywords):
                        continue

                archivos.append({
                    "url":    full_url,
                    "nombre": texto or href.split("/")[-1],
                    "ext":    ext,
                })

        # También buscar dentro de iframes (el SRI a veces embebe contenido)
        for iframe in soup.find_all("iframe", src=True):
            src = iframe["src"]
            if "docs.google.com" in src or "drive.google.com" in src:
                archivos.append({
                    "url":    src,
                    "nombre": "Google Sheet embebido",
                    "ext":    "gsheet",
                })

    except Exception as e:
        log.warning("Error buscando archivos en %s: %s", url, e)

    archivos.sort(key=lambda x: x["ext"] in ("xlsx", "csv", "xls"), reverse=True)
    return archivos


# ---------------------------------------------------------------------------
# CKAN — datosabiertos.gob.ec
# ---------------------------------------------------------------------------

def buscar_datasets_ckan(query: str, max_results: int = 5) -> list[dict]:
    """Busca datasets del SRI en el portal CKAN de datos abiertos."""
    recursos = []
    try:
        r = requests.get(
            f"{CKAN_API}/package_search",
            params={"q": query, "rows": max_results},
            headers=HEADERS,
            timeout=30,
        )
        data = r.json()
        if data.get("success"):
            for pkg in data["result"]["results"]:
                for rec in pkg.get("resources", []):
                    fmt = (rec.get("format") or "").upper()
                    if fmt in ("CSV", "XLS", "XLSX", "ODS", "JSON"):
                        recursos.append({
                            "nombre":  rec.get("name") or pkg.get("title", "sin_nombre"),
                            "url":     rec.get("url"),
                            "formato": fmt,
                            "paquete": pkg.get("name"),
                            "org":     pkg.get("organization", {}).get("title", ""),
                        })
    except Exception as e:
        log.warning("Error buscando en CKAN: %s", e)
    return recursos


# ---------------------------------------------------------------------------
# Scraping de tablas HTML del SRI
# ---------------------------------------------------------------------------

def scrape_tablas_html(url: str, tipo: str) -> list[dict]:
    """Extrae datos de tablas HTML en las páginas del SRI."""
    docs = []
    r = descargar(url)
    if not r:
        return docs

    soup = BeautifulSoup(r.text, "lxml")
    tablas = soup.find_all("table")

    for idx, tabla in enumerate(tablas):
        filas = tabla.find_all("tr")
        if len(filas) < 2:
            continue

        # Detectar cabeceras
        cabeceras = []
        for th in filas[0].find_all(["th", "td"]):
            texto = th.get_text(strip=True).replace("\n", " ")
            cabeceras.append(texto if texto else f"col_{len(cabeceras)}")

        if len(cabeceras) < 2:
            continue

        for fila in filas[1:]:
            celdas = fila.find_all(["td", "th"])
            if not celdas:
                continue
            valores = [c.get_text(strip=True) for c in celdas]
            if all(not v for v in valores):
                continue

            doc = {}
            for j, val in enumerate(valores):
                key = cabeceras[j] if j < len(cabeceras) else f"col_{j}"
                # Intentar convertir valores numéricos
                val_limpio = val.replace(",", "").replace(".", "").strip()
                if val_limpio.isdigit() and len(val_limpio) > 2:
                    try:
                        doc[key] = float(val.replace(",", ""))
                    except ValueError:
                        doc[key] = val
                else:
                    doc[key] = val

            doc["_tipo"]       = tipo
            doc["_fuente"]     = "SRI"
            doc["_tabla_idx"]  = idx
            doc["_url_origen"] = url
            doc["_ingestado"]  = datetime.now(timezone.utc)
            docs.append(doc)

    return docs


# ---------------------------------------------------------------------------
# Parseo Excel/CSV
# ---------------------------------------------------------------------------

def parsear_excel(contenido: bytes, nombre_seccion: str, url_origen: str) -> list[dict]:
    """Parsea archivo Excel del SRI con detección automática de cabeceras."""
    docs = []
    try:
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(contenido), read_only=True, data_only=True)

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            todas_filas = list(ws.iter_rows(values_only=True))
            if not todas_filas:
                continue

            cabeceras = None
            inicio_datos = 0
            for i, fila in enumerate(todas_filas):
                no_vacias = [c for c in fila if c is not None and str(c).strip()]
                if len(no_vacias) >= 3:
                    cabeceras = [
                        str(c).strip().replace("\n", " ") if c else f"col_{j}"
                        for j, c in enumerate(fila)
                    ]
                    inicio_datos = i + 1
                    break

            if not cabeceras:
                continue

            for fila in todas_filas[inicio_datos:]:
                if all(v is None for v in fila):
                    continue
                doc = dict(zip(cabeceras, fila))
                doc["_seccion"]    = nombre_seccion
                doc["_hoja"]       = sheet_name
                doc["_fuente"]     = "SRI"
                doc["_url_origen"] = url_origen
                doc["_ingestado"]  = datetime.now(timezone.utc)
                docs.append(doc)

    except ImportError:
        log.warning("openpyxl no disponible — guardando referencia de %s", nombre_seccion)
        docs.append({
            "_seccion":      nombre_seccion,
            "_fuente":       "SRI",
            "_tipo":         "excel_sin_parsear",
            "_tamaño_bytes": len(contenido),
            "_url_origen":   url_origen,
            "_ingestado":    datetime.now(timezone.utc),
        })
    except Exception as e:
        log.error("Error parseando Excel SRI %s: %s", nombre_seccion, e)

    return docs


def parsear_csv(contenido: bytes, nombre_seccion: str, url_origen: str) -> list[dict]:
    """Parsea CSV con detección de encoding."""
    docs = []
    for enc in ("utf-8-sig", "latin-1", "cp1252"):
        try:
            texto = contenido.decode(enc)
            reader = csv.DictReader(io.StringIO(texto))
            for row in reader:
                row["_seccion"]    = nombre_seccion
                row["_fuente"]     = "SRI"
                row["_url_origen"] = url_origen
                row["_ingestado"]  = datetime.now(timezone.utc)
                docs.append(row)
            break
        except Exception:
            continue
    return docs


# ---------------------------------------------------------------------------
# Driver por sección
# ---------------------------------------------------------------------------

def procesar_seccion(nombre: str, cfg: dict, db, modo_test: bool) -> int:
    col = db[cfg["col"]]
    col.create_index("_tipo")
    col.create_index("_ingestado")
    col.create_index("_seccion")

    total = 0

    # 1. Scrape tablas HTML
    for url in cfg.get("urls", []):
        log.info("  Scrapeando HTML: %s", url)
        docs = scrape_tablas_html(url, cfg["tipo"])
        if docs:
            if modo_test:
                docs = docs[:30]
            try:
                col.insert_many(docs, ordered=False)
                log.info("  HTML: %d docs de %s", len(docs), url)
                total += len(docs)
            except BulkWriteError as e:
                n = e.details.get("nInserted", 0)
                total += n
        time.sleep(2)

    # 2. Buscar y descargar archivos
    for url in cfg.get("urls", []):
        archivos = buscar_archivos_en_pagina(url)
        for archivo in archivos[:2 if modo_test else 8]:
            if archivo["ext"] in ("gsheet", "pdf"):
                # Google Sheets embebidos y PDFs: solo guardar referencia
                col.insert_one({
                    "_seccion":    nombre,
                    "_fuente":     "SRI",
                    "_tipo":       f"referencia_{archivo['ext']}",
                    "_nombre":     archivo["nombre"],
                    "_url_origen": archivo["url"],
                    "_ingestado":  datetime.now(timezone.utc),
                })
                total += 1
                continue

            log.info("  Descargando: %s (%s)", archivo["nombre"][:60], archivo["ext"])
            r = descargar(archivo["url"])
            if not r:
                continue

            if archivo["ext"] in ("xls", "xlsx"):
                docs = parsear_excel(r.content, nombre, archivo["url"])
            elif archivo["ext"] in ("csv", "ods"):
                docs = parsear_csv(r.content, nombre, archivo["url"])
            elif archivo["ext"] == "json":
                try:
                    data = json.loads(r.content)
                    docs = data if isinstance(data, list) else [data]
                    for d in docs:
                        d["_seccion"]    = nombre
                        d["_fuente"]     = "SRI"
                        d["_url_origen"] = archivo["url"]
                        d["_ingestado"]  = datetime.now(timezone.utc)
                except Exception:
                    docs = []
            else:
                docs = [{
                    "_seccion":      nombre,
                    "_fuente":       "SRI",
                    "_tipo":         f"archivo_{archivo['ext']}",
                    "_tamaño_bytes": len(r.content),
                    "_url_origen":   archivo["url"],
                    "_ingestado":    datetime.now(timezone.utc),
                }]

            if modo_test:
                docs = docs[:20]

            if docs:
                try:
                    col.insert_many(docs, ordered=False)
                    log.info("  %d docs de %s", len(docs), archivo["nombre"][:40])
                    total += len(docs)
                except BulkWriteError as e:
                    n = e.details.get("nInserted", 0)
                    total += n

            time.sleep(1)

    # 3. CKAN complementario
    ckan_query = cfg.get("ckan_query", "")
    if ckan_query:
        recursos = buscar_datasets_ckan(ckan_query, max_results=3)
        for rec in recursos:
            if not rec.get("url"):
                continue
            log.info("  CKAN SRI: %s (%s)", rec["nombre"][:50], rec["formato"])
            r = descargar(rec["url"])
            if not r:
                continue

            if rec["formato"] in ("CSV",):
                docs = parsear_csv(r.content, nombre, rec["url"])
            elif rec["formato"] in ("XLS", "XLSX"):
                docs = parsear_excel(r.content, nombre, rec["url"])
            elif rec["formato"] == "JSON":
                try:
                    data = json.loads(r.content)
                    docs = data if isinstance(data, list) else [data]
                    for d in docs:
                        d["_seccion"]    = nombre
                        d["_fuente"]     = "SRI"
                        d["_url_origen"] = rec["url"]
                        d["_ingestado"]  = datetime.now(timezone.utc)
                except Exception:
                    docs = []
            else:
                docs = []

            if docs:
                if modo_test:
                    docs = docs[:20]
                try:
                    col.insert_many(docs, ordered=False)
                    log.info("  CKAN: %d docs insertados", len(docs))
                    total += len(docs)
                except BulkWriteError as e:
                    total += e.details.get("nInserted", 0)

    return total


# ---------------------------------------------------------------------------
# Sync log
# ---------------------------------------------------------------------------

def registrar_sync_log(col_log, fuente: str, estado: str, detalle: dict):
    col_log.insert_one({
        "fuente":    fuente,
        "estado":    estado,
        "detalle":   detalle,
        "timestamp": datetime.now(timezone.utc),
    })


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main(secciones_filtro: list[str] = None, modo_test: bool = False):
    log.info("=== EcuaWatch · scraper_sri.py ===")

    client  = MongoClient(MONGO_URI)
    db      = client[DB_NAME]
    col_log = db[COL_SYNC_LOG]

    secciones = {
        k: v for k, v in SECCIONES_SRI.items()
        if not secciones_filtro or k in secciones_filtro
    }

    resumen = {}
    for nombre, cfg in secciones.items():
        log.info("Sección SRI: %s — %s", nombre, cfg["descripcion"])
        n = procesar_seccion(nombre, cfg, db, modo_test)
        resumen[nombre] = n

    total = sum(resumen.values())
    log.info("Finalizado SRI. Total: %d | %s", total, resumen)

    registrar_sync_log(
        col_log,
        fuente="sri",
        estado="completado",
        detalle={"resumen": resumen, "total": total},
    )
    client.close()


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="Scraper SRI — sri.gob.ec → MongoDB"
    )
    parser.add_argument(
        "--seccion",
        nargs="+",
        choices=list(SECCIONES_SRI.keys()),
        metavar="SECCION",
        help=f"Secciones a procesar: {', '.join(SECCIONES_SRI.keys())}",
    )
    parser.add_argument("--test", action="store_true", help="Modo prueba")
    args = parser.parse_args()
    main(secciones_filtro=args.seccion, modo_test=args.test)
