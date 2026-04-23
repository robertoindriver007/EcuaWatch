"""
scraper_judicial.py — Scraper de la Función Judicial del Ecuador
=================================================================
Portal:  https://www.funcionjudicial.gob.ec
         https://www.datosabiertos.gob.ec (CKAN — filtro Judicatura)
Datos:   Estadísticas judiciales y datos abiertos del Consejo de la Judicatura
           - Causas ingresadas por materia y provincia
           - Causas resueltas vs pendientes
           - Tiempos de resolución
           - Sentencias por tipo (civil, penal, laboral, etc.)
           - Personal judicial (jueces, fiscales)

Colección MongoDB: ecuador_intel.judicial.causas_estadisticas
                   ecuador_intel.judicial.sentencias
                   ecuador_intel.judicial.personal

IMPORTANCIA PARA ANÁLISIS CAUSAL:
  → Vincula procesos judiciales ↔ casos de corrupción detectados por Contraloría
  → Mide la impunidad: cuántas denuncias se archivan vs proceden
  → Correlaciona zonas de alta conflictividad ↔ pobreza (INEC) ↔ presupuesto (MEF)
  → Identifica cuellos de botella en la administración de justicia

Uso:
    python scraper_judicial.py               # Todo
    python scraper_judicial.py --test        # Modo prueba
"""

import argparse
import csv
import io
import json
import logging
import os
import sys
import time
from datetime import datetime, timezone
from typing import Optional
from urllib.parse import urljoin

import requests
from bs4 import BeautifulSoup
from pymongo import MongoClient
from pymongo.errors import BulkWriteError

# ---------------------------------------------------------------------------
# Configuración
# ---------------------------------------------------------------------------

MONGO_URI = os.environ.get(
    "MONGO_URI",
    "mongodb+srv://scraperbot:GJqMqljz4GYBT0PU@cluster0.nz7wcxv.mongodb.net/"
    "?retryWrites=true&w=majority&appName=Cluster0",
)
DB_NAME      = "ecuador_intel"
COL_SYNC_LOG = "_sync_log"

FJ_BASE  = "https://www.funcionjudicial.gob.ec"
CJ_BASE  = "https://www.consejodelajudicatura.gob.ec"
CKAN_API = "https://www.datosabiertos.gob.ec/api/3/action"

SECCIONES = {
    "causas_estadisticas": {
        "col":          "judicial.causas_estadisticas",
        "descripcion":  "Estadísticas de causas judiciales por materia, provincia y período",
        "ckan_queries": [
            "judicatura causas ingresadas",
            "función judicial estadísticas",
            "consejo judicatura causas",
        ],
        "paginas": [
            f"{FJ_BASE}/estadisticas",
            f"{FJ_BASE}/transparencia",
            f"{CJ_BASE}/transparencia",
        ],
    },
    "sentencias": {
        "col":          "judicial.sentencias",
        "descripcion":  "Sentencias y resoluciones judiciales por tipo y resultado",
        "ckan_queries": [
            "sentencias judiciales Ecuador",
            "resoluciones judiciales Judicatura",
        ],
        "paginas": [],
    },
    "personal": {
        "col":          "judicial.personal",
        "descripcion":  "Personal judicial: jueces, fiscales, defensores por provincia",
        "ckan_queries": [
            "jueces Ecuador personal judicial",
            "servidores judiciales",
        ],
        "paginas": [],
    },
}

HEADERS = {
    "User-Agent": "EcuaWatch-Bot/1.0 (Judicial research; ecuawatch.org)",
    "Accept": "text/html,application/json,application/xhtml+xml;q=0.9",
}

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[logging.StreamHandler(sys.stdout)],
)
log = logging.getLogger("scraper_judicial")

# ---------------------------------------------------------------------------
# HTTP + CKAN helpers
# ---------------------------------------------------------------------------

def descargar(url: str, timeout: int = 60) -> Optional[requests.Response]:
    for intento in range(3):
        try:
            r = requests.get(url, headers=HEADERS, timeout=timeout, allow_redirects=True)
            r.raise_for_status()
            return r
        except requests.RequestException as e:
            log.warning("Intento %d fallido para %s: %s", intento + 1, url, e)
            time.sleep(3 ** intento)
    return None


def buscar_ckan(query: str, max_results: int = 5) -> list[dict]:
    """Busca datasets del Consejo de la Judicatura en CKAN."""
    recursos = []
    try:
        r = requests.get(
            f"{CKAN_API}/package_search",
            params={"q": query, "rows": max_results},
            headers=HEADERS,
            timeout=30,
        )
        data = r.json()
        if data.get("success"):
            for pkg in data["result"]["results"]:
                for rec in pkg.get("resources", []):
                    fmt = (rec.get("format") or "").upper()
                    if fmt in ("CSV", "XLS", "XLSX", "JSON", "ODS"):
                        recursos.append({
                            "nombre":  rec.get("name") or pkg.get("title", ""),
                            "url":     rec.get("url"),
                            "formato": fmt,
                            "org":     pkg.get("organization", {}).get("title", ""),
                        })
    except Exception as e:
        log.warning("Error CKAN: %s", e)
    return recursos


def buscar_archivos_pagina(url: str) -> list[dict]:
    archivos = []
    try:
        r = descargar(url)
        if not r:
            return archivos
        soup = BeautifulSoup(r.text, "lxml")
        for a in soup.find_all("a", href=True):
            href = a["href"]
            if any(ext in href.lower() for ext in (".csv", ".xlsx", ".xls", ".json", ".ods")):
                archivos.append({
                    "url":  urljoin(url, href),
                    "nombre": a.get_text(strip=True) or href.split("/")[-1],
                    "ext":  href.split("?")[0].split(".")[-1].lower(),
                })
    except Exception as e:
        log.warning("Error en %s: %s", url, e)
    return archivos


def scrape_tablas_html(url: str, tipo: str) -> list[dict]:
    docs = []
    r = descargar(url)
    if not r:
        return docs
    soup = BeautifulSoup(r.text, "lxml")
    for idx, tabla in enumerate(soup.find_all("table")):
        filas = tabla.find_all("tr")
        if len(filas) < 2:
            continue
        cabeceras = [th.get_text(strip=True).replace("\n", " ") or f"col_{i}"
                     for i, th in enumerate(filas[0].find_all(["th", "td"]))]
        if len(cabeceras) < 2:
            continue
        for fila in filas[1:]:
            celdas = [c.get_text(strip=True) for c in fila.find_all(["td", "th"])]
            if all(not v for v in celdas):
                continue
            doc = {cabeceras[j] if j < len(cabeceras) else f"col_{j}": v
                   for j, v in enumerate(celdas)}
            doc.update({"_tipo": tipo, "_fuente": "CJ", "_tabla": idx,
                        "_url_origen": url, "_ingestado": datetime.now(timezone.utc)})
            docs.append(doc)
    return docs


# ---------------------------------------------------------------------------
# Parseo
# ---------------------------------------------------------------------------

def parsear_excel(contenido: bytes, seccion: str, url: str) -> list[dict]:
    docs = []
    try:
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(contenido), read_only=True, data_only=True)
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            filas = list(ws.iter_rows(values_only=True))
            if not filas:
                continue
            cabeceras = None
            inicio = 0
            for i, f in enumerate(filas):
                novacias = [c for c in f if c is not None and str(c).strip()]
                if len(novacias) >= 3:
                    cabeceras = [str(c).strip().replace("\n", " ") if c else f"col_{j}"
                                 for j, c in enumerate(f)]
                    inicio = i + 1
                    break
            if not cabeceras:
                continue
            for f in filas[inicio:]:
                if all(v is None for v in f):
                    continue
                doc = dict(zip(cabeceras, f))
                doc.update({"_seccion": seccion, "_hoja": sheet, "_fuente": "CJ",
                            "_url_origen": url, "_ingestado": datetime.now(timezone.utc)})
                docs.append(doc)
    except ImportError:
        docs.append({"_seccion": seccion, "_fuente": "CJ", "_tipo": "excel_ref",
                      "_tamaño": len(contenido), "_url_origen": url,
                      "_ingestado": datetime.now(timezone.utc)})
    except Exception as e:
        log.error("Error Excel %s: %s", seccion, e)
    return docs


def parsear_csv(contenido: bytes, seccion: str, url: str) -> list[dict]:
    docs = []
    for enc in ("utf-8-sig", "latin-1", "cp1252"):
        try:
            reader = csv.DictReader(io.StringIO(contenido.decode(enc)))
            for row in reader:
                row.update({"_seccion": seccion, "_fuente": "CJ",
                            "_url_origen": url, "_ingestado": datetime.now(timezone.utc)})
                docs.append(row)
            break
        except Exception:
            continue
    return docs


# ---------------------------------------------------------------------------
# Driver
# ---------------------------------------------------------------------------

def procesar_seccion(nombre: str, cfg: dict, db, modo_test: bool) -> int:
    col = db[cfg["col"]]
    col.create_index("_seccion")
    col.create_index("_ingestado")
    total = 0

    # 1. HTML tables
    for pagina in cfg.get("paginas", []):
        log.info("  HTML: %s", pagina)
        docs = scrape_tablas_html(pagina, nombre)
        if docs:
            if modo_test:
                docs = docs[:30]
            try:
                col.insert_many(docs, ordered=False)
                total += len(docs)
            except BulkWriteError as e:
                total += e.details.get("nInserted", 0)
        time.sleep(2)

    # 2. Archivos en páginas
    for pagina in cfg.get("paginas", []):
        for archivo in buscar_archivos_pagina(pagina)[:5]:
            log.info("  Archivo: %s (%s)", archivo["nombre"][:50], archivo["ext"])
            r = descargar(archivo["url"])
            if not r:
                continue
            if archivo["ext"] in ("xls", "xlsx"):
                docs = parsear_excel(r.content, nombre, archivo["url"])
            elif archivo["ext"] == "csv":
                docs = parsear_csv(r.content, nombre, archivo["url"])
            else:
                docs = [{"_seccion": nombre, "_fuente": "CJ",
                          "_tipo": f"ref_{archivo['ext']}", "_url_origen": archivo["url"],
                          "_ingestado": datetime.now(timezone.utc)}]
            if modo_test:
                docs = docs[:20]
            if docs:
                try:
                    col.insert_many(docs, ordered=False)
                    total += len(docs)
                except BulkWriteError as e:
                    total += e.details.get("nInserted", 0)
            time.sleep(1)

    # 3. CKAN
    for query in cfg.get("ckan_queries", []):
        for rec in buscar_ckan(query, max_results=3):
            if not rec.get("url"):
                continue
            log.info("  CKAN: %s (%s)", rec["nombre"][:50], rec["formato"])
            r = descargar(rec["url"])
            if not r:
                continue
            if rec["formato"] in ("CSV",):
                docs = parsear_csv(r.content, nombre, rec["url"])
            elif rec["formato"] in ("XLS", "XLSX"):
                docs = parsear_excel(r.content, nombre, rec["url"])
            elif rec["formato"] == "JSON":
                try:
                    data = json.loads(r.content)
                    docs = data if isinstance(data, list) else [data]
                    for d in docs:
                        d.update({"_seccion": nombre, "_fuente": "CJ",
                                  "_url_origen": rec["url"],
                                  "_ingestado": datetime.now(timezone.utc)})
                except Exception:
                    docs = []
            else:
                docs = []
            if docs:
                if modo_test:
                    docs = docs[:20]
                try:
                    col.insert_many(docs, ordered=False)
                    total += len(docs)
                except BulkWriteError as e:
                    total += e.details.get("nInserted", 0)

    return total


def main(modo_test: bool = False):
    log.info("=== EcuaWatch · scraper_judicial.py ===")
    client  = MongoClient(MONGO_URI)
    db      = client[DB_NAME]
    col_log = db[COL_SYNC_LOG]

    resumen = {}
    for nombre, cfg in SECCIONES.items():
        log.info("Sección Judicial: %s — %s", nombre, cfg["descripcion"])
        n = procesar_seccion(nombre, cfg, db, modo_test)
        resumen[nombre] = n

    total = sum(resumen.values())
    log.info("Finalizado Judicial. Total: %d | %s", total, resumen)
    col_log.insert_one({"fuente": "judicial", "estado": "completado",
                         "detalle": {"resumen": resumen, "total": total},
                         "timestamp": datetime.now(timezone.utc)})
    client.close()


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Scraper Judicial → MongoDB")
    parser.add_argument("--test", action="store_true")
    args = parser.parse_args()
    main(modo_test=args.test)
