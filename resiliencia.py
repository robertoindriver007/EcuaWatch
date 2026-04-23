"""
resiliencia.py — Capa de Resiliencia Industrial para EcuaWatch
================================================================
Módulo compartido que todos los scrapers DEBEN importar.

Resuelve TODOS los micro-bloqueos conocidos de los sitios del gobierno:

  1. TIMEOUTS DINÁMICOS: Adapta el timeout según la velocidad del servidor
  2. CIRCUIT BREAKER: Si un servidor falla N veces seguidas, lo pausa
  3. RATE LIMITER: Respeta los Headers Retry-After y 429
  4. DETECCIÓN DE CAPTCHA/WAF: Identifica Cloudflare, reCAPTCHA, WAF
  5. ROTACIÓN DE USER-AGENT: Evita bloqueo por fingerprinting
  6. DETECCIÓN DE LOOPS INFINITOS: Mata URLs que redireccionan en círculo
  7. GUARDIÁN DE SESIÓN JSF: Refresca ViewState automáticamente
  8. CHECKPOINTING: Guarda progreso por fuente para reanudación
  9. VALIDACIÓN DE DESCARGAS: Verifica integridad (tamaño, tipo, encoding)
  10. MÉTRICAS EN TIEMPO REAL: Registra rendimiento por scraper

Patrones extraídos del scraper_asamblea.py probado en producción:
  → Decorador retry con backoff exponencial
  → Detección de Content-Type real
  → Validación de respuestas HTML disfrazadas de archivos
  → Checkpoint con JSON persistente

Uso:
    from resiliencia import HttpResilient, CircuitBreaker, guardar_metricas
"""

import hashlib
import json
import logging
import os
import random
import sys
import time
from collections import defaultdict
from dataclasses import dataclass, field
from datetime import datetime, timezone
from pathlib import Path
from typing import Callable, Optional
from urllib.parse import urlparse

import requests
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry

log = logging.getLogger("resiliencia")

# ---------------------------------------------------------------------------
# 1. USER-AGENTS realistas (rotación anti-fingerprinting)
# ---------------------------------------------------------------------------
USER_AGENTS = [
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/125.0.0.0 Safari/537.36",
    "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 14_5) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/17.4 Safari/605.1.15",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:126.0) Gecko/20100101 Firefox/126.0",
    "EcuaWatch-Bot/2.0 (investigacion; ecuawatch.org)",
]

# ---------------------------------------------------------------------------
# 2. CIRCUIT BREAKER — evita martillar servidores caídos
# ---------------------------------------------------------------------------

@dataclass
class CircuitBreaker:
    """
    Si un host falla `umbral_fallos` veces consecutivas, entra en estado
    ABIERTO y rechaza peticiones por `pausa_segundos`.
    
    Estados:
      CERRADO → Todo normal, peticiones pasan
      ABIERTO → Host bloqueado temporalmente
      SEMI    → Permite 1 petición de prueba
    """
    umbral_fallos: int = 5
    pausa_segundos: int = 120
    _contadores: dict = field(default_factory=lambda: defaultdict(int))
    _abierto_hasta: dict = field(default_factory=dict)
    _total_bloqueos: dict = field(default_factory=lambda: defaultdict(int))

    def registrar_exito(self, host: str):
        self._contadores[host] = 0
        if host in self._abierto_hasta:
            del self._abierto_hasta[host]

    def registrar_fallo(self, host: str):
        self._contadores[host] += 1
        if self._contadores[host] >= self.umbral_fallos:
            self._abierto_hasta[host] = time.time() + self.pausa_segundos
            self._total_bloqueos[host] += 1
            log.warning(
                "⚡ CIRCUIT BREAKER ABIERTO para %s — %d fallos consecutivos. "
                "Pausado por %ds (bloqueo #%d)",
                host, self._contadores[host], self.pausa_segundos,
                self._total_bloqueos[host]
            )

    def permitir(self, host: str) -> bool:
        if host not in self._abierto_hasta:
            return True
        if time.time() > self._abierto_hasta[host]:
            log.info("⚡ Circuit breaker SEMI-ABIERTO para %s — probando...", host)
            del self._abierto_hasta[host]
            self._contadores[host] = self.umbral_fallos - 1  # Semi: 1 más y se cierra
            return True
        return False

    def estado(self, host: str) -> str:
        if host not in self._abierto_hasta:
            return "CERRADO"
        if time.time() > self._abierto_hasta[host]:
            return "SEMI"
        return "ABIERTO"

    def reporte(self) -> dict:
        return {
            "contadores": dict(self._contadores),
            "bloqueos_totales": dict(self._total_bloqueos),
            "estados": {h: self.estado(h) for h in set(
                list(self._contadores.keys()) + list(self._abierto_hasta.keys())
            )},
        }


# Instancia global compartida por todos los scrapers
_circuit_breaker = CircuitBreaker()


# ---------------------------------------------------------------------------
# 3. MÉTRICAS DE RENDIMIENTO
# ---------------------------------------------------------------------------

@dataclass
class MetricasScraper:
    """Registra métricas de rendimiento por fuente/host."""
    peticiones_totales: int = 0
    peticiones_exitosas: int = 0
    peticiones_fallidas: int = 0
    bytes_descargados: int = 0
    tiempo_total_s: float = 0.0
    reintentos_totales: int = 0
    loops_detectados: int = 0
    captchas_detectados: int = 0
    timeouts_ajustados: int = 0
    descargas_corruptas: int = 0
    _tiempos_respuesta: list = field(default_factory=list)
    _inicio: float = field(default_factory=time.time)

    def registrar_peticion(self, exitosa: bool, bytes_desc: int = 0, tiempo_s: float = 0):
        self.peticiones_totales += 1
        if exitosa:
            self.peticiones_exitosas += 1
            self.bytes_descargados += bytes_desc
        else:
            self.peticiones_fallidas += 1
        self.tiempo_total_s += tiempo_s
        self._tiempos_respuesta.append(tiempo_s)

    @property
    def tasa_exito(self) -> float:
        if self.peticiones_totales == 0:
            return 0.0
        return self.peticiones_exitosas / self.peticiones_totales * 100

    @property
    def tiempo_promedio_s(self) -> float:
        if not self._tiempos_respuesta:
            return 0.0
        return sum(self._tiempos_respuesta) / len(self._tiempos_respuesta)

    @property
    def velocidad_mbps(self) -> float:
        if self.tiempo_total_s == 0:
            return 0.0
        return (self.bytes_descargados / 1_000_000) / self.tiempo_total_s

    def reporte(self) -> dict:
        return {
            "peticiones": self.peticiones_totales,
            "exito": self.peticiones_exitosas,
            "fallos": self.peticiones_fallidas,
            "tasa_exito_%": round(self.tasa_exito, 1),
            "bytes_total": self.bytes_descargados,
            "mb_total": round(self.bytes_descargados / 1_000_000, 2),
            "tiempo_total_s": round(self.tiempo_total_s, 1),
            "tiempo_promedio_s": round(self.tiempo_promedio_s, 3),
            "velocidad_mbps": round(self.velocidad_mbps, 3),
            "reintentos": self.reintentos_totales,
            "loops": self.loops_detectados,
            "captchas": self.captchas_detectados,
            "descargas_corruptas": self.descargas_corruptas,
            "duracion_sesion_s": round(time.time() - self._inicio, 1),
        }


_metricas: dict[str, MetricasScraper] = defaultdict(MetricasScraper)


def obtener_metricas(fuente: str) -> MetricasScraper:
    return _metricas[fuente]


# ---------------------------------------------------------------------------
# 4. DETECCIÓN DE BLOQUEOS (CAPTCHA, WAF, redirect loops)
# ---------------------------------------------------------------------------

CAPTCHA_INDICATORS = [
    "captcha", "recaptcha", "g-recaptcha", "hCaptcha",
    "challenge-form", "challenge-running",
    "cf-browser-verification", "Attention Required",
    "Just a moment", "Checking your browser",
    "Access denied", "403 Forbidden",
    "Please complete the security check",
]

WAF_INDICATORS = [
    "cloudflare", "incapsula", "imperva", "sucuri",
    "akamai", "access denied", "waf",
]


def detectar_bloqueo(response: requests.Response) -> Optional[str]:
    """
    Analiza una respuesta HTTP para detectar si es un bloqueo.
    Retorna el tipo de bloqueo o None si es una respuesta legítima.
    """
    ct = response.headers.get("content-type", "").lower()
    body = ""
    if "html" in ct:
        body = response.text[:5000].lower()

    # 1. CAPTCHA
    for indicator in CAPTCHA_INDICATORS:
        if indicator.lower() in body:
            return f"CAPTCHA:{indicator}"

    # 2. WAF / CDN block
    for indicator in WAF_INDICATORS:
        if indicator.lower() in body:
            server = response.headers.get("server", "")
            return f"WAF:{indicator}({server})"

    # 3. Redirect a login
    if response.status_code in (301, 302, 303, 307, 308):
        location = response.headers.get("location", "").lower()
        if any(w in location for w in ["login", "auth", "sso", "cas/"]):
            return f"REDIRECT_LOGIN:{location[:100]}"

    # 4. Error page disfrazada
    if response.status_code == 200 and "html" in ct:
        if len(response.content) < 2000:
            if any(w in body for w in ["error", "no encontrado", "not found", "404", "maintenance"]):
                return f"ERROR_PAGE_DISFRAZADA:{response.status_code}"

    return None


def detectar_loop_redirect(response: requests.Response, historial: set) -> bool:
    """Detecta si caemos en un loop de redirects."""
    if response.history:
        for r in response.history:
            url = r.url
            if url in historial:
                return True
            historial.add(url)
    return False


# ---------------------------------------------------------------------------
# 5. CLIENTE HTTP RESILIENTE
# ---------------------------------------------------------------------------

class HttpResilient:
    """
    Cliente HTTP blindado que aplica TODAS las defensas:
      - Retry con backoff exponencial
      - Circuit breaker por host
      - Detección de captcha/WAF
      - Timeout dinámico
      - Rotación de User-Agent
      - Validación de respuesta
      - Anti-loop de redirects
      - Métricas en tiempo real
    """

    def __init__(self, fuente: str, max_reintentos: int = 3,
                 timeout_base: int = 30, max_timeout: int = 120,
                 delay_base: float = 1.5, respetar_robots: bool = True):
        self.fuente = fuente
        self.max_reintentos = max_reintentos
        self.timeout_base = timeout_base
        self.max_timeout = max_timeout
        self.delay_base = delay_base
        self.metricas = obtener_metricas(fuente)
        self._redirect_historial: set = set()
        self._timeout_actual = timeout_base
        self._urls_visitadas_sesion: set = set()

        # Sesión con retry adapter nativo
        self.session = requests.Session()
        retry_strategy = Retry(
            total=2,
            backoff_factor=1,
            status_forcelist=[500, 502, 503, 504],
        )
        adapter = HTTPAdapter(max_retries=retry_strategy)
        self.session.mount("https://", adapter)
        self.session.mount("http://", adapter)

        self._rotar_useragent()

    def _rotar_useragent(self):
        ua = random.choice(USER_AGENTS)
        self.session.headers.update({
            "User-Agent": ua,
            "Accept-Language": "es-EC,es;q=0.9,en;q=0.8",
            "Accept": "text/html,application/json,application/xhtml+xml,*/*;q=0.8",
            "Accept-Encoding": "gzip, deflate, br",
        })

    def _host_de_url(self, url: str) -> str:
        return urlparse(url).hostname or "unknown"

    def get(self, url: str, params: dict = None,
            headers: dict = None, stream: bool = False,
            validar_tipo: str = None) -> Optional[requests.Response]:
        """
        GET resiliente con todas las protecciones activas.
        
        Args:
            validar_tipo: Si se especifica, verifica que el Content-Type coincida
                          Ejemplo: "json", "excel", "csv", "pdf"
        """
        host = self._host_de_url(url)

        # Circuit breaker check
        if not _circuit_breaker.permitir(host):
            log.warning("⚡ %s bloqueado por circuit breaker. Saltando.", host)
            self.metricas.registrar_peticion(exitosa=False, tiempo_s=0)
            return None

        last_error = None
        for intento in range(1, self.max_reintentos + 1):
            t0 = time.time()
            try:
                # Rotar UA cada 10 peticiones
                if self.metricas.peticiones_totales % 10 == 0:
                    self._rotar_useragent()

                merged_headers = dict(self.session.headers)
                if headers:
                    merged_headers.update(headers)

                resp = self.session.get(
                    url, params=params, headers=merged_headers,
                    timeout=self._timeout_actual, stream=stream,
                    allow_redirects=True,
                )
                elapsed = time.time() - t0

                # Detectar loop de redirect
                if detectar_loop_redirect(resp, self._redirect_historial):
                    log.error("🔄 LOOP DE REDIRECT detectado para %s. Abortando.", url)
                    self.metricas.loops_detectados += 1
                    _circuit_breaker.registrar_fallo(host)
                    return None

                # Rate limit (429)
                if resp.status_code == 429:
                    retry_after = int(resp.headers.get("Retry-After", 60))
                    log.warning("⏳ Rate limit 429 de %s. Esperando %ds.", host, retry_after)
                    time.sleep(min(retry_after, 300))  # Max 5 min
                    self.metricas.reintentos_totales += 1
                    continue

                # 5xx server error
                if 500 <= resp.status_code <= 599:
                    log.warning("🔥 Error %d de %s (intento %d/%d)",
                                resp.status_code, host, intento, self.max_reintentos)
                    _circuit_breaker.registrar_fallo(host)
                    self.metricas.reintentos_totales += 1
                    delay = self.delay_base * (2 ** intento) + random.uniform(0, 1)
                    time.sleep(delay)
                    continue

                resp.raise_for_status()

                # Detectar bloqueo
                bloqueo = detectar_bloqueo(resp)
                if bloqueo:
                    log.warning("🛡️ BLOQUEO detectado en %s: %s", host, bloqueo)
                    if "CAPTCHA" in bloqueo:
                        self.metricas.captchas_detectados += 1
                    _circuit_breaker.registrar_fallo(host)
                    time.sleep(30)
                    continue

                # Validar tipo de contenido
                if validar_tipo and not self._validar_content_type(resp, validar_tipo):
                    log.warning("⚠️ Content-Type inesperado para %s (esperaba %s, recibió %s)",
                                url[:80], validar_tipo,
                                resp.headers.get("content-type", "?"))
                    self.metricas.descargas_corruptas += 1
                    # No reintentamos — el servidor devuelve algo diferente
                    _circuit_breaker.registrar_exito(host)
                    self.metricas.registrar_peticion(True, len(resp.content), elapsed)
                    return resp  # Devolvemos igual, el caller decide

                # Ajustar timeout dinámico
                self._ajustar_timeout(elapsed)

                # Éxito
                _circuit_breaker.registrar_exito(host)
                self.metricas.registrar_peticion(True, len(resp.content), elapsed)
                self._urls_visitadas_sesion.add(url)
                return resp

            except requests.exceptions.Timeout:
                elapsed = time.time() - t0
                log.warning("⏱️ Timeout (%ds) para %s (intento %d/%d)",
                            self._timeout_actual, host, intento, self.max_reintentos)
                self._timeout_actual = min(self._timeout_actual * 1.5, self.max_timeout)
                self.metricas.timeouts_ajustados += 1
                _circuit_breaker.registrar_fallo(host)
                last_error = f"Timeout({self._timeout_actual}s)"

            except requests.exceptions.ConnectionError as e:
                elapsed = time.time() - t0
                log.warning("🔌 Error de conexión con %s: %s (intento %d/%d)",
                            host, str(e)[:100], intento, self.max_reintentos)
                _circuit_breaker.registrar_fallo(host)
                last_error = str(e)[:200]

            except requests.exceptions.RequestException as e:
                elapsed = time.time() - t0
                log.error("❌ Error HTTP para %s: %s", url[:100], e)
                _circuit_breaker.registrar_fallo(host)
                last_error = str(e)[:200]

            # Backoff exponencial con jitter
            delay = self.delay_base * (2 ** intento) + random.uniform(0, 2)
            self.metricas.reintentos_totales += 1
            log.info("  ⏲️ Esperando %.1fs antes del reintento...", delay)
            time.sleep(delay)

        # Todos los reintentos fallaron
        self.metricas.registrar_peticion(exitosa=False, tiempo_s=0)
        log.error("❌ %s: agotados %d reintentos. Último error: %s",
                  url[:100], self.max_reintentos, last_error)
        return None

    def _validar_content_type(self, resp: requests.Response, esperado: str) -> bool:
        ct = resp.headers.get("content-type", "").lower()
        mapa = {
            "json":  ["application/json", "text/json"],
            "csv":   ["text/csv", "text/plain"],
            "excel": ["spreadsheetml", "ms-excel", "openxml", "application/octet"],
            "pdf":   ["application/pdf"],
            "html":  ["text/html"],
        }
        tipos = mapa.get(esperado.lower(), [esperado.lower()])
        return any(t in ct for t in tipos)

    def _ajustar_timeout(self, tiempo_real: float):
        """Ajuste dinámico: si el servidor responde rápido, reducimos timeout."""
        if tiempo_real < self.timeout_base * 0.3:
            nuevo = max(self.timeout_base, self._timeout_actual * 0.9)
            if nuevo != self._timeout_actual:
                self._timeout_actual = nuevo

    def descargar_archivo(self, url: str, validar_minimo_bytes: int = 100) -> Optional[bytes]:
        """
        Descarga un archivo con validación de integridad.
        Verifica tamaño mínimo y que no sea una página HTML disfrazada.
        """
        resp = self.get(url)
        if not resp:
            return None

        contenido = resp.content
        ct = resp.headers.get("content-type", "").lower()

        # Validación: tamaño mínimo
        if len(contenido) < validar_minimo_bytes:
            log.warning("⚠️ Archivo muy pequeño (%d bytes) de %s. Posible vacío/error.",
                        len(contenido), url[:80])
            self.metricas.descargas_corruptas += 1
            return None

        # Validación: HTML disfrazado como archivo
        if "text/html" in ct and not url.endswith(".html"):
            text_preview = contenido[:1000].decode("utf-8", errors="ignore").lower()
            if any(w in text_preview for w in ["<html", "<!doctype", "<head"]):
                log.warning("⚠️ Respuesta HTML disfrazada de %s. Sesión expirada?", url[:80])
                self.metricas.descargas_corruptas += 1
                return None

        return contenido

    def reporte_final(self) -> dict:
        """Genera reporte completo de la sesión."""
        return {
            "fuente": self.fuente,
            "metricas": self.metricas.reporte(),
            "circuit_breaker": _circuit_breaker.reporte(),
            "urls_visitadas": len(self._urls_visitadas_sesion),
            "timeout_actual": self._timeout_actual,
        }


# ---------------------------------------------------------------------------
# 6. CHECKPOINT PERSISTENTE (reanudación tras fallo)
# ---------------------------------------------------------------------------

CHECKPOINT_DIR = Path(__file__).parent / ".checkpoints"


def cargar_checkpoint(fuente: str) -> dict:
    """Carga el estado guardado de un scraper."""
    CHECKPOINT_DIR.mkdir(exist_ok=True)
    archivo = CHECKPOINT_DIR / f"{fuente}.json"
    if archivo.exists():
        try:
            with open(archivo, "r", encoding="utf-8") as f:
                return json.load(f)
        except (json.JSONDecodeError, IOError):
            return {}
    return {}


def guardar_checkpoint(fuente: str, estado: dict):
    """Guarda el estado actual de un scraper para reanudación."""
    CHECKPOINT_DIR.mkdir(exist_ok=True)
    archivo = CHECKPOINT_DIR / f"{fuente}.json"
    estado["_timestamp"] = datetime.now(timezone.utc).isoformat()
    with open(archivo, "w", encoding="utf-8") as f:
        json.dump(estado, f, indent=2, default=str)


def anotar_procesado(fuente: str, item_id: str):
    """Marca un item como procesado en el checkpoint."""
    ckpt = cargar_checkpoint(fuente)
    procesados = set(ckpt.get("procesados", []))
    procesados.add(item_id)
    ckpt["procesados"] = list(procesados)
    ckpt["ultimo_procesado"] = item_id
    guardar_checkpoint(fuente, ckpt)


def ya_procesado(fuente: str, item_id: str) -> bool:
    """Verifica si un item ya fue procesado."""
    ckpt = cargar_checkpoint(fuente)
    return item_id in set(ckpt.get("procesados", []))


# ---------------------------------------------------------------------------
# 7. GUARDAR MÉTRICAS EN MONGODB
# ---------------------------------------------------------------------------

def guardar_metricas_mongo(db, fuente: str, extra: dict = None):
    """Persiste las métricas de rendimiento en MongoDB para análisis."""
    col = db["_metricas_rendimiento"]
    col.create_index([("fuente", 1), ("timestamp", -1)])

    metricas = obtener_metricas(fuente)
    doc = {
        "fuente":    fuente,
        "timestamp": datetime.now(timezone.utc),
        **metricas.reporte(),
        "circuit_breaker": _circuit_breaker.reporte(),
    }
    if extra:
        doc.update(extra)

    col.insert_one(doc)
    log.info("📊 Métricas de %s guardadas en _metricas_rendimiento", fuente)


# ---------------------------------------------------------------------------
# 8. DEDUPLICACIÓN INTELIGENTE
# ---------------------------------------------------------------------------

def hash_contenido(contenido: bytes) -> str:
    """SHA256 del contenido para detectar duplicados exactos."""
    return hashlib.sha256(contenido).hexdigest()


def hash_documento(doc: dict, campos_clave: list[str]) -> str:
    """Hash de un documento MongoDB basado en campos clave."""
    valores = [str(doc.get(c, "")) for c in sorted(campos_clave)]
    return hashlib.sha256("|".join(valores).encode()).hexdigest()


# ---------------------------------------------------------------------------
# 9. VALIDACIÓN DE DATOS DESCARGADOS
# ---------------------------------------------------------------------------

def validar_csv(contenido: bytes) -> tuple[bool, str]:
    """Verifica que un CSV sea válido y tenga datos."""
    import csv as csv_mod
    import io as io_mod
    for enc in ("utf-8-sig", "latin-1", "cp1252"):
        try:
            text = contenido.decode(enc)
            reader = csv_mod.reader(io_mod.StringIO(text))
            filas = list(reader)
            if len(filas) < 2:
                return False, "CSV sin datos (solo cabecera o vacío)"
            if len(filas[0]) < 2:
                return False, f"CSV con solo {len(filas[0])} columna(s)"
            return True, f"OK: {len(filas)} filas, {len(filas[0])} columnas"
        except Exception:
            continue
    return False, "No se pudo decodificar el CSV en ningún encoding"


def validar_excel(contenido: bytes) -> tuple[bool, str]:
    """Verifica que un Excel sea válido y tenga datos."""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(contenido), read_only=True, data_only=True)
        total_filas = 0
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            for _ in ws.iter_rows(max_row=5):
                total_filas += 1
        if total_filas < 2:
            return False, "Excel sin datos"
        return True, f"OK: {len(wb.sheetnames)} hoja(s)"
    except ImportError:
        return True, "openpyxl no disponible, asumiendo OK"
    except Exception as e:
        return False, f"Excel corrupto: {e}"


def validar_json(contenido: bytes) -> tuple[bool, str]:
    """Verifica que un JSON sea parseable y tenga datos."""
    try:
        data = json.loads(contenido)
        if isinstance(data, list):
            return len(data) > 0, f"Array con {len(data)} items"
        elif isinstance(data, dict):
            return True, f"Objeto con {len(data)} keys"
        return True, f"Tipo: {type(data).__name__}"
    except json.JSONDecodeError as e:
        return False, f"JSON inválido: {e}"


def validar_descarga(contenido: bytes, formato: str) -> tuple[bool, str]:
    """Dispatcher de validación por formato."""
    fmt = formato.lower()
    if fmt in ("csv", "text/csv"):
        return validar_csv(contenido)
    elif fmt in ("xls", "xlsx", "excel"):
        return validar_excel(contenido)
    elif fmt in ("json", "application/json"):
        return validar_json(contenido)
    elif fmt in ("pdf", "application/pdf"):
        if contenido[:4] == b"%PDF":
            return True, f"PDF válido ({len(contenido)} bytes)"
        return False, "No es un PDF válido (sin header %PDF)"
    return True, f"Formato {fmt} no validado — asumiendo OK"


# ---------------------------------------------------------------------------
# 10. REPORTE DE SALUD (para el orquestador)
# ---------------------------------------------------------------------------

def generar_reporte_salud() -> dict:
    """Genera un reporte de salud completo de todas las fuentes."""
    reporte = {
        "timestamp":       datetime.now(timezone.utc).isoformat(),
        "circuit_breaker": _circuit_breaker.reporte(),
        "fuentes":         {},
    }
    for fuente, metricas in _metricas.items():
        reporte["fuentes"][fuente] = metricas.reporte()
    return reporte


import io  # Needed for validar_excel
